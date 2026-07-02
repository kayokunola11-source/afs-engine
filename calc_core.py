# -*- coding: utf-8 -*-
"""
AFS Studio — calc core (Slice 1 seed).

Computes the trial balance, P&L subtotals and the Nigerian tax charge for a
standard-template workbook *in pure Python, from the INPUT sheets only*
(COA, Cover, OpenBal, JE, Cashbook, BankImport, TB_Import).

It does NOT read the workbook's computed TB/statement/CIT sheets, and it does
NOT use LibreOffice. Those computed sheets are used only by `validate()` as a
regression oracle to prove this engine reproduces them.

Design rules (from the architecture blueprint):
  * one source of truth for logic (this module)
  * fail loud, never silent — unmapped codes and out-of-balance TBs are reported
  * map to statements by COA *section*, not by label text or code prefix
  * account codes are normalised once, here, because the workbook stores some as
    text ("9100") and some as numbers (1700.0) and Excel silently coerces them.

Validated 2026-07-01 against BODY TEMPLE_2025_FIXED.xlsx (98 accounts, 4,940
bank lines, 17 journals): TB matches to <5 kobo; every CIT line matches exactly.

Usage:
    python calc_core.py "BODY TEMPLE_2025_FIXED.xlsx"     # prints a tie-out report
"""
from __future__ import annotations
import sys
from collections import defaultdict
import openpyxl


# ----------------------------------------------------------------- helpers
def ncode(v):
    """Canonical account-code string. 1700.0 -> '1700', '9100' -> '9100'."""
    if v is None or isinstance(v, bool):
        return ""
    if isinstance(v, float):
        return str(int(v)) if v.is_integer() else str(v)
    if isinstance(v, int):
        return str(v)
    s = str(v).strip()
    return s[:-2] if (s.endswith(".0") and s[:-2].isdigit()) else s


def num(v):
    """Numeric value or 0.0. Text like '-' (first-year suppressor) becomes 0."""
    return float(v) if isinstance(v, (int, float)) else 0.0


# ----------------------------------------------------------------- inputs
def read_inputs(wb):
    """Read the input sheets into plain Python structures. Returns a dict."""
    coa = {}
    ws = wb["COA"]
    for r in range(5, 113):
        c = ncode(ws.cell(r, 1).value)
        if not c:
            continue
        coa[c] = {
            "name": ws.cell(r, 2).value,
            "cls": str(ws.cell(r, 3).value or "").strip(),
            "section": str(ws.cell(r, 4).value or "").strip(),
            "normal": str(ws.cell(r, 5).value or "").strip(),
        }

    opening = {}
    ws = wb["OpenBal"]
    for r in range(5, ws.max_row + 1):
        c = ncode(ws.cell(r, 1).value)
        if not c or c.upper() == "TOTAL":
            continue
        opening[c] = {
            "dr": num(ws.cell(r, 3).value), "cr": num(ws.cell(r, 4).value),
            "pydr": num(ws.cell(r, 5).value), "pycr": num(ws.cell(r, 6).value),
        }

    # movements: (debit_code, credit_code, amount) from every input source
    postings = []  # list of (dr_code, cr_code, amount)

    # Journals: each row is a single-sided posting keyed by its own code
    ws = wb["JE"]
    for r in range(5, 1001):
        c = ncode(ws.cell(r, 3).value)
        if not c:
            continue
        dr, cr = num(ws.cell(r, 5).value), num(ws.cell(r, 6).value)
        if dr:
            postings.append((c, None, dr))
        if cr:
            postings.append((None, c, cr))

    # Cashbook: In -> Dr bank / Cr counter ; Out -> Cr bank / Dr counter
    ws = wb["Cashbook"]
    for r in range(5, 2001):
        bank = ncode(ws.cell(r, 2).value); cnt = ncode(ws.cell(r, 4).value)
        inv, outv = num(ws.cell(r, 5).value), num(ws.cell(r, 6).value)
        if inv:
            postings.append((bank or None, cnt or None, inv))
        if outv:
            postings.append((cnt or None, bank or None, outv))

    # BankImport: In -> Dr bank(A) / Cr counter(H) ; Out -> Cr bank / Dr counter
    ws = wb["BankImport"]
    for r in range(5, ws.max_row + 1):
        bank = ncode(ws.cell(r, 1).value); cnt = ncode(ws.cell(r, 8).value)
        inv, outv = num(ws.cell(r, 5).value), num(ws.cell(r, 4).value)
        if inv:
            postings.append((bank or None, cnt or None, inv))
        if outv:
            postings.append((cnt or None, bank or None, outv))

    # TB_Import (mapped Dr/Cr helper cols J/K keyed by firm code F; asset-manager
    # workbooks instead carry a single "Mapped Net (Dr-Cr)" in col H — fall back to it)
    ws = wb["TB_Import"]
    for r in range(8, 1001):
        c = ncode(ws.cell(r, 6).value)
        if not c:
            continue
        dr, cr = num(ws.cell(r, 10).value), num(ws.cell(r, 11).value)
        if not dr and not cr:
            net = num(ws.cell(r, 8).value)
            if net > 0:
                dr = net
            elif net < 0:
                cr = -net
        if dr:
            postings.append((c, None, dr))
        if cr:
            postings.append((None, c, cr))

    cover = _read_cover(wb["Cover"])
    return {"coa": coa, "opening": opening, "postings": postings, "cover": cover}


def _read_cover(ws):
    v = lambda r: ws.cell(r, 3).value
    return {
        "entity": v(5), "rc": v(6), "period_end": v(7), "prior_end": v(8),
        "first_year": str(v(18) or "No").strip().lower().startswith("y"),
        "cit_rate": num(v(20)), "tet_rate": num(v(21)), "dev_levy": num(v(22)),
        "min_tax": num(v(23)), "ptf_rate": num(v(53)),
        "small_ceiling": num(v(51)), "medium_ceiling": num(v(52)),
        "size": str(v(55) or "").strip(),
    }


# ----------------------------------------------------------------- trial balance
def build_trial_balance(inp):
    """Return (tb, errors). tb[code] = dict of OB/movement/closing figures."""
    coa, opening, postings = inp["coa"], inp["opening"], inp["postings"]
    mdr, mcr = defaultdict(float), defaultdict(float)
    errors = []
    for dr_code, cr_code, amt in postings:
        if dr_code:
            if dr_code not in coa:
                errors.append(f"posting debit to unknown code {dr_code!r}")
            mdr[dr_code] += amt
        if cr_code:
            if cr_code not in coa:
                errors.append(f"posting credit to unknown code {cr_code!r}")
            mcr[cr_code] += amt

    tb = {}
    for code in sorted(set(list(coa) + list(opening) + list(mdr) + list(mcr))):
        ob = opening.get(code, {})
        obdr, obcr = num(ob.get("dr")), num(ob.get("cr"))
        g, h = mdr.get(code, 0.0), mcr.get(code, 0.0)
        net = obdr - obcr + g - h
        tb[code] = {
            "name": coa.get(code, {}).get("name"),
            "section": coa.get(code, {}).get("section", ""),
            "cls": coa.get(code, {}).get("cls", ""),
            "obdr": obdr, "obcr": obcr, "mdr": g, "mcr": h,
            "cydr": net if net > 0 else 0.0, "cycr": -net if net < 0 else 0.0,
            "cy_signed": net,  # Dr positive
            "pydr": num(ob.get("pydr")), "pycr": num(ob.get("pycr")),
        }

    tot_dr = sum(a["cydr"] for a in tb.values())
    tot_cr = sum(a["cycr"] for a in tb.values())
    if abs(tot_dr - tot_cr) > 1.0:
        errors.append(f"trial balance out of balance by {tot_dr - tot_cr:,.2f}")
    return tb, errors


# ----------------------------------------------------------------- statements
def statement_subtotals(tb):
    """P&L subtotals by COA section (Dr positive)."""
    s = defaultdict(float)
    for a in tb.values():
        if a["section"].startswith("PL"):
            s[a["section"]] += a["cy_signed"]
    revenue = -s["PL-Revenue"]
    other_income = -(s["PL-OtherInc"] + s["PL-OtherGains"])
    cos, admin, selling, finance = s["PL-COS"], s["PL-Admin"], s["PL-Selling"], s["PL-FinCost"]
    gross = revenue - cos
    pbt = revenue + other_income - cos - admin - selling - finance
    return {
        "revenue": revenue, "other_income": other_income, "cost_of_sales": cos,
        "gross_profit": gross, "admin_expenses": admin, "selling_expenses": selling,
        "finance_cost": finance, "profit_before_tax": pbt,
        "tax_booked": s["PL-Tax"],
    }


# ----------------------------------------------------------------- tax
def compute_tax(tb, pbt, cover, cap_allow_cy=0.0):
    """Nigerian company tax charge, mirroring the CIT schedule (regime-aware)."""
    small = cover["size"].lower() == "small"
    revenue = -sum(a["cy_signed"] for a in tb.values() if a["section"] == "PL-Revenue")
    dep = tb.get("6800", {}).get("cy_signed", 0.0) + tb.get("6810", {}).get("cy_signed", 0.0)
    donations = tb.get("6420", {}).get("cy_signed", 0.0)

    adjusted = pbt + dep + donations
    assessable = adjusted  # + balancing charge (0 by default)
    cap = -min(cap_allow_cy, max(0.0, (2 / 3) * assessable))
    taxable = assessable + cap

    cit = 0.0 if small else (0.0 if taxable <= 0 else max(taxable * cover["cit_rate"], revenue * cover["min_tax"]))
    tet = 0.0 if small else (0.0 if assessable <= 0 else max(0.0, assessable * cover["tet_rate"]))
    dev = 0.0 if small else max(0.0, assessable * cover["dev_levy"])
    ptf = 0.0 if small else max(0.0, pbt) * cover["ptf_rate"]
    return {
        "profit_before_tax": pbt, "depreciation_addback": dep,
        "adjusted_profit": adjusted, "assessable_profit": assessable,
        "capital_allowance": cap, "taxable_income": taxable,
        "cit": cit, "tet": tet, "development_levy": dev, "police_trust_fund": ptf,
        "total_tax": cit + tet + dev + ptf,
    }


def compute(path):
    """Full pipeline for one workbook. Returns a structured result dict."""
    wb = openpyxl.load_workbook(path, data_only=True)
    inp = read_inputs(wb)
    tb, errors = build_trial_balance(inp)
    subs = statement_subtotals(tb)
    cap_cy = num(wb["CapAllow"]["H13"].value) if "CapAllow" in wb.sheetnames else 0.0
    tax = compute_tax(tb, subs["profit_before_tax"], inp["cover"], cap_cy)
    return {"cover": inp["cover"], "tb": tb, "subtotals": subs, "tax": tax, "errors": errors}


# ----------------------------------------------------------------- regression oracle
def validate(path):
    """Compare the pure-Python result to the workbook's own computed cells."""
    wb = openpyxl.load_workbook(path, data_only=True)
    res = compute(path)
    g = lambda cell, sheet="CIT": num(wb[sheet][cell].value)

    checks = [
        ("Revenue",             res["subtotals"]["revenue"],           g("D6", "SOCI")),
        ("Gross profit",        res["subtotals"]["gross_profit"],      g("D8", "SOCI")),
        ("Profit before tax",   res["subtotals"]["profit_before_tax"], g("D15", "SOCI")),
        ("CIT adjusted profit", res["tax"]["adjusted_profit"],         g("C16")),
        ("CIT taxable income",  res["tax"]["taxable_income"],          g("C21")),
        ("Companies Income Tax",res["tax"]["cit"],                     g("C23")),
        ("Tertiary Ed. Tax",    res["tax"]["tet"],                     g("C24")),
        ("Development Levy",     res["tax"]["development_levy"],        g("C25")),
        ("Police Trust Fund",   res["tax"]["police_trust_fund"],       g("C26")),
        ("Total tax charge",    res["tax"]["total_tax"],               g("C27")),
    ]
    print(f"\nTie-out vs workbook  —  {path}")
    print(f"  entity: {res['cover']['entity']}   size: {res['cover']['size']}")
    print("  " + "-" * 60)
    all_ok = True
    for label, py, truth in checks:
        ok = abs(py - truth) < 0.10
        all_ok &= ok
        print(f"  [{'OK ' if ok else 'DIFF'}] {label:22s} {py:16,.2f}  vs {truth:16,.2f}")
    tot_dr = sum(a["cydr"] for a in res["tb"].values())
    tot_cr = sum(a["cycr"] for a in res["tb"].values())
    print("  " + "-" * 60)
    print(f"  TB totals: Dr {tot_dr:,.2f}  Cr {tot_cr:,.2f}  diff {tot_dr - tot_cr:,.2f}")
    if res["errors"]:
        print("  FLAGS (fail-loud):")
        for e in res["errors"][:10]:
            print("    -", e)
    print(f"\n  RESULT: {'ALL TIE-OUTS PASS' if all_ok else 'MISMATCH — investigate'}")
    return all_ok


def selfcheck(path):
    """Production-safe cross-check: run the pure-Python engine on a workbook and
    compare to its own computed cells. Returns a JSON-friendly dict. Never raises
    for accounting reasons — callers should still wrap in try/except for safety."""
    wb = openpyxl.load_workbook(path, data_only=True)
    res = compute(path)
    g = lambda cell, sheet="CIT": num(wb[sheet][cell].value)
    pairs = [
        ("revenue",            res["subtotals"]["revenue"],           g("D6", "SOCI")),
        ("profit_before_tax",  res["subtotals"]["profit_before_tax"], g("D15", "SOCI")),
        ("taxable_income",     res["tax"]["taxable_income"],          g("C21")),
        ("companies_income_tax", res["tax"]["cit"],                   g("C23")),
        ("total_tax_charge",   res["tax"]["total_tax"],              g("C27")),
    ]
    checks = [{"name": n, "python": round(py, 2), "workbook": round(tr, 2),
               "match": abs(py - tr) < 0.10} for n, py, tr in pairs]
    tot_dr = sum(a["cydr"] for a in res["tb"].values())
    tot_cr = sum(a["cycr"] for a in res["tb"].values())
    return {
        "engine": "calc_core-slice1",
        "entity": res["cover"]["entity"],
        "size": res["cover"]["size"],
        "all_match": all(c["match"] for c in checks),
        "tb_out_of_balance": round(tot_dr - tot_cr, 2),
        "checks": checks,
        "flags": res["errors"][:10],
    }


if __name__ == "__main__":
    path = sys.argv[1] if len(sys.argv) > 1 else "BODY TEMPLE_2025_FIXED.xlsx"
    validate(path)
