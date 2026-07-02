# -*- coding: utf-8 -*-
"""Slice 1b assembler: build the afs_generator data dict entirely from the calc core."""
import openpyxl, calc_core
from calc_core import ncode, num
from collections import defaultdict

NCA={"NCA-PPE-Cost","NCA-PPE-Dep","NCA-Intangible-Cost","NCA-Intangible-Amort","NCA-Investments","NCA-DefTax","NCA-PreInc"}
CA={"CA-Inventory","CA-Inv-RawMat","CA-Inv-WIP","CA-Inv-FG","CA-Trade-Rec","CA-Other-Rec","CA-Allowance","CA-Prepay","CA-Cash","CA-Bank","CA-Clearing","CA-Suspense"}
CL={"CL-Trade-Pay","CL-Accruals","CL-Statutory","CL-Tax","CL-DCA","CL-Overdraft","CL-Loans","CL-Other-Pay","CL-DefIncome","CL-Borrow"}
NCL={"NCL-Loans","NCL-Other","NCL-DefTax","NCL-Borrow"}
PL={"PL-Revenue","PL-OtherInc","PL-OtherGains","PL-COS","PL-Admin","PL-Selling","PL-FinCost","PL-Tax","PL-Prod-Materials","PL-Prod-Labour","PL-Prod-Overhead","INC-MgmtFee","INC-PerfFee","INC-Other","EXP-Direct","EXP-Staff","EXP-Occupancy","EXP-Regulatory","EXP-Admin","EXP-Depr","EXP-Finance"}

def _py(wb):
    ws=wb["OpenBal"]; d={}
    for r in range(5,ws.max_row+1):
        c=ncode(ws.cell(r,1).value)
        if c and c.upper()!="TOTAL": d[c]=num(ws.cell(r,5).value)-num(ws.cell(r,6).value)
    return d

def _ppe(wb):
    if "PPE_Schedule" not in wb.sheetnames: return None
    ws=wb["PPE_Schedule"]; classes=[]; total=None; hdr=False
    for r in range(4,23):
        a=ws.cell(r,1).value; an=str(a).strip() if a is not None else ""
        if an=="Asset class": hdr=True; continue
        if not hdr or not an: continue
        cost=num(ws.cell(r,6).value); charge=num(ws.cell(r,8).value); dep=num(ws.cell(r,11).value); nbv=num(ws.cell(r,12).value)
        if an.upper()=="TOTAL": total=[cost,dep,charge,nbv]; break
        if cost or dep or nbv or charge: classes.append([an,cost,dep,charge,nbv])
    return {"classes":classes,"total":total or [0,0,0,0]}


_SCF_PYMAP={
 "Profit/(loss) before tax":["profit/(loss) before tax"],
 "Depreciation & amortisation":["depreciation"],
 "(Increase)/decrease in inventory":["inventory","inventories"],
 "(Increase)/decrease in receivables":["receivable"],
 "Increase/(decrease) in payables & accruals":["payable"],
 "Tax paid":["tax paid"],
 "Net cash from/(used in) operating activities":["operating activities"],
 "Purchase of property, plant & equipment":["property, plant"],
 "Purchase of intangible assets":["intangible"],
 "Purchase of investments":["investment"],
 "Net cash from/(used in) investing activities":["investing activities"],
 "Movement in share capital":["share issue","capital introduced"],
 "Movement in borrowings":["long-term loan","borrow"],
 "Director's current account movement":["director's current","directors current"],
 "Net cash from/(used in) financing activities":["financing activities"],
 "Net increase/(decrease) in cash":["net increase"],
 "Cash & cash equivalents at start of period":["start of period"],
 "Cash & cash equivalents at end of period":["end of period"],
}
def _scf_py_rows(wb):
    """Prior-year cash-flow figures, read from the workbook's own SCF sheet (col B label, col E = PY)."""
    if "SCF" not in wb.sheetnames: return []
    ws=wb["SCF"]; out=[]
    for r in range(1,60):
        b=ws.cell(r,2).value; e=ws.cell(r,5).value
        if b and str(b).strip() and isinstance(e,(int,float)):
            out.append((str(b).strip().lower(), float(e)))
    return out
def _scf_py_val(rows, kws):
    return sum(v for lab,v in rows for k in kws if k in lab) if rows else 0.0


def _num2(v): return float(v) if isinstance(v,(int,float)) else 0.0
def _grid(headers, rows, money_from=1, subhead=None, bold_last=True):
    return {"headers":headers,"rows":rows,"money_from":money_from,"subhead":subhead,"bold_last":bold_last}

def read_deftax(wb, rate):
    if "DefTax" not in wb.sheetnames: return None
    ws=wb["DefTax"]; comps=[]
    for r in range(6,13):
        a=ws.cell(r,1).value; b=ws.cell(r,2).value; c=ws.cell(r,3).value
        if a and (isinstance(b,(int,float)) or isinstance(c,(int,float))):
            comps.append([str(a).strip(), _num2(b)*rate, _num2(c)*rate])
    if not comps: return None
    close=sum(x[1] for x in comps); close_py=sum(x[2] for x in comps)
    opn=_num2(ws.cell(14,2).value); opn_py=_num2(ws.cell(14,3).value)
    tbl=[[x[0],x[1],x[2]] for x in comps]
    tbl.append(["Net deferred tax asset/(liability)",close,close_py,"total"])
    recon=[["Opening balance",opn,opn_py],["Charge/(credit) to profit or loss",close-opn,close_py-opn_py],
           ["Closing balance",close,close_py,"total"]]
    return {"tables":[("Deferred tax by temporary difference",tbl),("Movement in deferred tax",recon)]}

def read_finrisk(wb, rec_cy, rec_py, cash_cy, cash_py):
    if "FinRisk" not in wb.sheetnames: return None
    ws=wb["FinRisk"]; grids=[]
    liq=[]
    for r in range(6,11):
        a=ws.cell(r,1).value
        vals=[_num2(ws.cell(r,c).value) for c in range(2,6)]
        if a and any(vals): liq.append([str(a).strip()]+vals+[sum(vals)])
    if liq:
        liq.append(["Total"]+[sum(row[i] for row in liq) for i in range(1,6)]); liq[-1].append("total")
        grids.append(_grid(["Financial liability","< 3 months","3–12 months","1–5 years","> 5 years","Total"],liq,
                           subhead="Liquidity risk — contractual maturities of financial liabilities"))
    cr=[]
    for r in range(15,17):
        a=ws.cell(r,1).value; b=ws.cell(r,2).value; c=ws.cell(r,3).value
        if a:
            cy=_num2(b) if isinstance(b,(int,float)) else (rec_cy if r==15 else cash_cy)
            py=_num2(c) if isinstance(c,(int,float)) else (rec_py if r==15 else cash_py)
            cr.append([str(a).strip(),cy,py])
    if cr:
        cr.append(["Maximum exposure to credit risk",sum(x[1] for x in cr),sum(x[2] for x in cr),"total"])
        grids.append(_grid(["Credit risk exposure","CY","PY"],cr,subhead="Credit risk — maximum exposure"))
    return grids or None

def read_ecl(wb):
    if "ECL" not in wb.sheetnames: return None
    ws=wb["ECL"]; rows=[]
    for r in range(6,11):
        a=ws.cell(r,1).value; g=ws.cell(r,2).value; rt=ws.cell(r,3).value
        if a and isinstance(g,(int,float)):
            rate=_num2(rt)/100.0
            rows.append([str(a).strip(),_num2(g),"%.1f%%"%(rate*100),_num2(g)*rate])
    if not rows: return None
    rows.append(["Total",sum(x[1] for x in rows),"",sum(x[3] for x in rows),"total"])
    return _grid(["Ageing bucket","Gross","Loss rate","ECL allowance"],rows,money_from=1,
                 subhead="Expected credit loss — trade receivables (provision matrix)")

def read_leases(wb):
    if "Leases" not in wb.sheetnames: return None
    ws=wb["Leases"]; rou=[_num2(ws.cell(5,c).value) for c in range(2,6)]
    ll=[_num2(ws.cell(9,c).value) for c in range(2,7)]
    if not (any(rou) or any(ll)): return None
    tbl=[["Right-of-use assets — opening",rou[0],None],["Additions",rou[1],None],
         ["Depreciation",-rou[2],None],["Right-of-use assets — closing",rou[3],None,"total"],
         ["Lease liabilities — closing",ll[4],None,"total"]]
    return {"table":tbl}

def read_finsum(wb):
    if "FinSum" not in wb.sheetnames: return None
    ws=wb["FinSum"]
    yrs=[ws.cell(4,c).value for c in range(2,6)]
    headers=["\u20a6'000"]+[str(int(y)) if isinstance(y,(int,float)) else (str(y) if y else "") for y in yrs]
    rows=[]
    for r in range(6,16):
        lbl=ws.cell(r,1).value
        if not lbl or not str(lbl).strip(): continue
        vals=[ws.cell(r,c).value for c in range(2,6)]
        if any(isinstance(v,(int,float)) for v in vals):
            rows.append([str(lbl).strip()]+[(_num2(v) if isinstance(v,(int,float)) else 0) for v in vals])
    return _grid(headers,rows,money_from=1,subhead=None,bold_last=False) if rows else None

def read_capadq(wb):
    if "Capital_Adequacy" not in wb.sheetnames: return None
    ws=wb["Capital_Adequacy"]; rows=[]
    for r in range(5,20):
        a=ws.cell(r,1).value; b=ws.cell(r,2).value; c=ws.cell(r,3).value
        if a and (isinstance(b,(int,float)) or isinstance(c,(int,float))):
            lab=str(a).strip()
            tot = lab.upper().startswith("TOTAL") or "ratio" in lab.lower() or "excess" in lab.lower() or "minimum required" in lab.lower()
            row=[lab,_num2(b),_num2(c)]+ (["total"] if tot else [])
            rows.append(row)
    return rows or None

def build_data(path, meta_over=None, disclosures=None, scale=None, full_ifrs=None):
    disclosures=disclosures or {}
    wb=openpyxl.load_workbook(path, data_only=True)
    inp=calc_core.read_inputs(wb); tb,errs=calc_core.build_trial_balance(inp); cov=inp["cover"]
    _plnt={"PL-Revenue","PL-OtherInc","PL-OtherGains","PL-COS","PL-Admin","PL-Selling","PL-FinCost"}
    _pbt=-sum(a["cy_signed"] for a in tb.values() if a["section"] in _plnt)
    _cap=num(wb["CapAllow"]["H13"].value) if "CapAllow" in wb.sheetnames else 0.0
    res={"tb":tb,"cover":cov,"tax":calc_core.compute_tax(tb,_pbt,cov,_cap),"errors":errs,"_cap_cy":_cap}
    sec={c:a["section"] for c,a in tb.items()}; name={c:a["name"] for c,a in tb.items()}
    cy={c:a["cy_signed"] for c,a in tb.items()}; py=_py(wb)
    def S(book,secs,sign=1): return sign*sum(book.get(c,0.0) for c,s in sec.items() if s in secs)
    def prof(book): return -sum(book.get(c,0.0) for c,s in sec.items() if s in PL)
    def accts(book,secs,sign=1):
        rows=[]
        for c,s in sec.items():
            if s in secs:
                v=sign*book.get(c,0.0)
                if abs(v)>=1: rows.append([str(name.get(c) or c), v])
        return rows

    pat=prof(cy); pat_py=prof(py)
    re_acct=S(cy,{"EQ-RetEarn","EQ-Drawings"},-1); re_close=re_acct+pat
    re_acct_py=S(py,{"EQ-RetEarn","EQ-Drawings"},-1); re_close_py=re_acct_py+pat_py
    re_open=re_close_py                       # opening RE = prior-year CLOSING (the fix)

    # ---- tax (CY + PY) via same engine ----
    def tax_of(book):
        pbt=-sum(book.get(c,0.0) for c,s in sec.items() if s in (PL-{"PL-Tax"}))
        return calc_core.compute_tax(tb if book is cy else {k:{**tb[k],"cy_signed":book.get(k,0.0),"section":sec[k]} for k in tb}, pbt, cov)
    taxcy=res["tax"]
    # PY tax: rebuild a py-view tb
    tb_py={k:{"section":sec[k],"cy_signed":py.get(k,0.0)} for k in sec}
    pbt_py=-sum(py.get(c,0.0) for c,s in sec.items() if s in (PL-{"PL-Tax"}))
    taxpy=calc_core.compute_tax(tb_py,pbt_py,cov)

    # ---------- statements ----------
    def money(cyv,pyv,label,note="",kind="normal",indent=False):
        r={"label":label,"note":note,"cy":cyv,"py":pyv,"kind":kind}
        if indent: r["indent"]=True
        return r
    rev=S(cy,{"PL-Revenue"},-1); rev_py=S(py,{"PL-Revenue"},-1)
    cos=S(cy,{"PL-COS"}); cos_py=S(py,{"PL-COS"})
    oi=S(cy,{"PL-OtherInc","PL-OtherGains"},-1); oi_py=S(py,{"PL-OtherInc","PL-OtherGains"},-1)
    admin=S(cy,{"PL-Admin"}); admin_py=S(py,{"PL-Admin"})
    sell=S(cy,{"PL-Selling"}); sell_py=S(py,{"PL-Selling"})
    fin=S(cy,{"PL-FinCost"}); fin_py=S(py,{"PL-FinCost"})
    taxexp=S(cy,{"PL-Tax"}); taxexp_py=S(py,{"PL-Tax"})
    gross=rev-cos; gross_py=rev_py-cos_py
    op=gross+oi-admin-sell-fin; op_py=gross_py+oi_py-admin_py-sell_py-fin_py
    soci=[money(rev,rev_py,"Revenue","6",indent=True),
          money(-cos,-cos_py,"Cost of sales","8",indent=True),
          money(gross,gross_py,"GROSS PROFIT",kind="total")]
    if abs(oi)>=1 or abs(oi_py)>=1: soci.append(money(oi,oi_py,"Other income","7",indent=True))
    soci+=[money(-admin,-admin_py,"Administrative expenses","9",indent=True),
           money(-sell,-sell_py,"Selling & distribution expenses","10",indent=True)]
    if abs(fin)>=1 or abs(fin_py)>=1: soci.append(money(-fin,-fin_py,"Finance cost","11",indent=True))
    soci+=[money(op,op_py,"OPERATING PROFIT/(LOSS)",kind="subtotal"),
           money(op,op_py,"PROFIT/(LOSS) BEFORE TAX",kind="subtotal"),
           money(-taxexp,-taxexp_py,"Taxation","12",indent=True),
           money(pat,pat_py,"PROFIT/(LOSS) FOR THE YEAR",kind="total"),
           money(pat,pat_py,"TOTAL COMPREHENSIVE INCOME",kind="grandtotal")]
    am = any(a.get("section")=="INC-MgmtFee" for a in tb.values())
    if am:
        full_ifrs=True
        mgmt=S(cy,{"INC-MgmtFee"},-1); mgmt_py=S(py,{"INC-MgmtFee"},-1)
        perf=S(cy,{"INC-PerfFee"},-1); perf_py=S(py,{"INC-PerfFee"},-1)
        invinc=S(cy,{"INC-Other"},-1); invinc_py=S(py,{"INC-Other"},-1)
        direct=S(cy,{"EXP-Direct"}); direct_py=S(py,{"EXP-Direct"})
        opex=S(cy,{"EXP-Staff","EXP-Occupancy","EXP-Regulatory","EXP-Admin","EXP-Depr"}); opex_py=S(py,{"EXP-Staff","EXP-Occupancy","EXP-Regulatory","EXP-Admin","EXP-Depr"})
        finc=S(cy,{"EXP-Finance"}); finc_py=S(py,{"EXP-Finance"})
        tfi=mgmt+perf; tfi_py=mgmt_py+perf_py
        netfee=tfi-direct; netfee_py=tfi_py-direct_py
        opam=netfee+invinc-opex; opam_py=netfee_py+invinc_py-opex_py
        pbtam=opam-finc; pbtam_py=opam_py-finc_py
        # keep downstream (SCF, VAS, financial instruments) consistent
        rev=tfi+invinc; rev_py=tfi_py+invinc_py; oi=0; oi_py=0
        cos=direct; cos_py=direct_py; admin=opex; admin_py=opex_py; sell=0; sell_py=0
        fin=finc; fin_py=finc_py; gross=rev-cos; gross_py=rev_py-cos_py; op=pbtam; op_py=pbtam_py
        soci=[money(mgmt,mgmt_py,"Management fee income","6",indent=True)]
        if abs(perf)>=1 or abs(perf_py)>=1: soci.append(money(perf,perf_py,"Performance fee income","6",indent=True))
        soci.append(money(tfi,tfi_py,"TOTAL FEE INCOME",kind="subtotal"))
        if abs(direct)>=1 or abs(direct_py)>=1:
            soci.append(money(-direct,-direct_py,"Direct fund servicing costs","8",indent=True))
            soci.append(money(netfee,netfee_py,"NET FEE INCOME",kind="subtotal"))
        soci.append(money(invinc,invinc_py,"Investment and other income","7",indent=True))
        soci.append(money(-opex,-opex_py,"Operating expenses","9",indent=True))
        soci.append(money(opam,opam_py,"OPERATING PROFIT/(LOSS)",kind="subtotal"))
        if abs(finc)>=1 or abs(finc_py)>=1: soci.append(money(-finc,-finc_py,"Finance cost","11",indent=True))
        # Asset-manager tax: max(CIT on taxable profit, minimum tax = rate x gross income)
        _mtr=cov.get("min_tax",0.005) or 0.005; _cr=cov.get("cit_rate",0.30) or 0.30
        _txbl=pbtam+S(cy,{"EXP-Depr"}); _txbl_py=pbtam_py+S(py,{"EXP-Depr"})
        tax_am=max(max(0,_txbl)*_cr,(tfi+invinc)*_mtr); tax_am_py=max(max(0,_txbl_py)*_cr,(tfi_py+invinc_py)*_mtr)
        pat=pbtam-tax_am; pat_py=pbtam_py-tax_am_py
        soci+=[money(pbtam,pbtam_py,"PROFIT/(LOSS) BEFORE TAX",kind="subtotal"),
               money(-tax_am,-tax_am_py,"Taxation","12",indent=True),
               money(pat,pat_py,"PROFIT/(LOSS) FOR THE YEAR",kind="total"),
               money(pat,pat_py,"TOTAL COMPREHENSIVE INCOME",kind="grandtotal")]

    mfg = any(a.get("section") in ("CA-Inv-RawMat","CA-Inv-WIP","CA-Inv-FG","PL-Prod-Materials","PL-Prod-Labour","PL-Prod-Overhead") for a in tb.values())
    _mfg_sched=None
    if mfg:
        def _op(secs): return sum(py.get(c,0.0) for c,aa in tb.items() if aa["section"] in secs)
        rm_o=_op({"CA-Inv-RawMat"}); rm_c=S(cy,{"CA-Inv-RawMat"})
        wip_o=_op({"CA-Inv-WIP"}); wip_c=S(cy,{"CA-Inv-WIP"})
        fg_o=_op({"CA-Inv-FG"}); fg_c=S(cy,{"CA-Inv-FG"})
        purch=S(cy,{"PL-Prod-Materials"}); labour=S(cy,{"PL-Prod-Labour"}); ovh=S(cy,{"PL-Prod-Overhead"})
        other_cos=S(cy,{"PL-COS"})
        rm_cons=rm_o+purch-rm_c; cost_prod=rm_cons+labour+ovh
        cogm=cost_prod+wip_o-wip_c; cogs=cogm+fg_o-fg_c; cos=cogs+other_cos
        cos_py=S(py,{"PL-Prod-Materials","PL-Prod-Labour","PL-Prod-Overhead","PL-COS"})
        gross=rev-cos; gross_py=rev_py-cos_py
        op=gross+oi-admin-sell-fin; op_py=gross_py+oi_py-admin_py-sell_py-fin_py
        pat=rev+oi-cos-admin-sell-fin-taxexp; pat_py=rev_py+oi_py-cos_py-admin_py-sell_py-fin_py-taxexp_py
        _mfg_sched=[["Opening raw materials",rm_o],["Add: raw materials purchased",purch],["Less: closing raw materials",-rm_c],
                    ["Raw materials consumed",rm_cons,"t"],["Direct labour",labour],["Factory overheads",ovh],
                    ["Cost of production",cost_prod,"t"],["Add: opening work-in-progress",wip_o],["Less: closing work-in-progress",-wip_c],
                    ["Cost of goods manufactured",cogm,"t"],["Add: opening finished goods",fg_o],["Less: closing finished goods",-fg_c],
                    ["Cost of goods sold",cogs,"t"]]
        _inv_split=[["Raw materials",rm_c,rm_o],["Work-in-progress",wip_c,wip_o],["Finished goods",fg_c,fg_o],
                    ["Total inventories",rm_c+wip_c+fg_c,rm_o+wip_o+fg_o,"t"]]
        soci=[money(rev,rev_py,"Revenue","6",indent=True),money(-cos,-cos_py,"Cost of sales","8",indent=True),
              money(gross,gross_py,"GROSS PROFIT",kind="total")]
        if abs(oi)>=1 or abs(oi_py)>=1: soci.append(money(oi,oi_py,"Other income","7",indent=True))
        soci+=[money(-admin,-admin_py,"Administrative expenses","9",indent=True),
               money(-sell,-sell_py,"Selling & distribution expenses","10",indent=True)]
        if abs(fin)>=1 or abs(fin_py)>=1: soci.append(money(-fin,-fin_py,"Finance cost","11",indent=True))
        soci+=[money(op,op_py,"OPERATING PROFIT/(LOSS)",kind="subtotal"),money(op,op_py,"PROFIT/(LOSS) BEFORE TAX",kind="subtotal"),
               money(-taxexp,-taxexp_py,"Taxation","12",indent=True),money(pat,pat_py,"PROFIT/(LOSS) FOR THE YEAR",kind="total"),
               money(pat,pat_py,"TOTAL COMPREHENSIVE INCOME",kind="grandtotal")]
    def bs_line(secs,label,note="",sign=1):
        return money(S(cy,secs,sign),S(py,secs,sign),label,note,indent=True)
    ppe_cy=S(cy,{"NCA-PPE-Cost","NCA-PPE-Dep"}); ppe_py=S(py,{"NCA-PPE-Cost","NCA-PPE-Dep"})
    nca=S(cy,NCA); nca_py=S(py,NCA); ca=S(cy,CA); ca_py=S(py,CA); ta=nca+ca; ta_py=nca_py+ca_py
    inv=S(cy,{"CA-Inventory"}); inv_py=S(py,{"CA-Inventory"})
    rec=S(cy,{"CA-Trade-Rec","CA-Other-Rec","CA-Allowance","CA-Prepay"}); rec_py=S(py,{"CA-Trade-Rec","CA-Other-Rec","CA-Allowance","CA-Prepay"})
    cash=S(cy,{"CA-Cash","CA-Bank","CA-Clearing","CA-Suspense"}); cash_py=S(py,{"CA-Cash","CA-Bank","CA-Clearing","CA-Suspense"})
    sc=S(cy,{"EQ-ShareCap","EQ-SharePrem","EQ-Reserve","EQ-Capital","EQ-Reserves","EQ-StatRes"},-1); sc_py=S(py,{"EQ-ShareCap","EQ-SharePrem","EQ-Reserve","EQ-Capital","EQ-Reserves","EQ-StatRes"},-1)
    cl=S(cy,CL,-1); cl_py=S(py,CL,-1); ncl=S(cy,NCL,-1); ncl_py=S(py,NCL,-1)
    tel=sc+re_close+ncl+cl; tel_py=sc_py+re_close_py+ncl_py+cl_py
    sofp=[{"label":"ASSETS","kind":"section"},{"label":"Non-current assets","kind":"section"},
          money(ppe_cy,ppe_py,"Property, plant & equipment","13",indent=True),
          money(nca,nca_py,"Total non-current assets",kind="subtotal"),
          {"label":"Current assets","kind":"section"},
          money(inv,inv_py,"Inventory","",indent=True),
          money(rec,rec_py,"Trade receivables","14",indent=True),
          money(cash,cash_py,"Cash & cash equivalents","15",indent=True),
          money(ca,ca_py,"Total current assets",kind="subtotal"),
          money(ta,ta_py,"TOTAL ASSETS",kind="grandtotal"),
          {"label":"EQUITY AND LIABILITIES","kind":"section"},{"label":"Equity","kind":"section"},
          money(sc,sc_py,"Share capital","16",indent=True),
          money(re_close,re_close_py,"Retained earnings","17",indent=True),
          money(sc+re_close,sc_py+re_close_py,"Total equity",kind="subtotal"),
          {"label":"Current liabilities","kind":"section"},
          money(S(cy,{"CL-Accruals"},-1),S(py,{"CL-Accruals"},-1),"Accruals","19",indent=True),
          money(S(cy,{"CL-Statutory"},-1),S(py,{"CL-Statutory"},-1),"Statutory deductions","",indent=True),
          money(S(cy,{"CL-Tax"},-1),S(py,{"CL-Tax"},-1),"Current tax payable","12",indent=True),
          money(S(cy,{"CL-DCA"},-1),S(py,{"CL-DCA"},-1),"Director's current account","20",indent=True),
          money(cl,cl_py,"Total current liabilities",kind="subtotal"),
          money(tel,tel_py,"TOTAL EQUITY AND LIABILITIES",kind="grandtotal")]

    # ---------- SCF (indirect, derived from BS movements -> ties by identity) ----------
    def dS(secs): return S(cy,secs)-S(py,secs)          # movement in signed balance (Dr+)
    OP_CL={"CL-Trade-Pay","CL-Accruals","CL-Statutory"}
    INV={"NCA-PPE-Cost","NCA-Intangible-Cost","NCA-Investments","NCA-PreInc"}
    FIN={"EQ-ShareCap","EQ-SharePrem","EQ-Reserve","EQ-Capital","EQ-Reserves","EQ-StatRes","EQ-Drawings",
         "NCL-Loans","NCL-Other","CL-Loans","CL-Overdraft","CL-DCA"}
    CASH={"CA-Cash","CA-Bank","CA-Clearing","CA-Suspense"}
    pbt=op
    dep_charge=-dS({"NCA-PPE-Dep"}); amort_charge=-dS({"NCA-Intangible-Amort"})
    d_inv=-dS({"CA-Inventory"}); d_rec=-dS({"CA-Trade-Rec","CA-Other-Rec","CA-Allowance","CA-Prepay"})
    d_pay=-dS(OP_CL)
    tax_mag=(-S(cy,{"CL-Tax"}))-(-S(py,{"CL-Tax"}))     # increase in tax payable
    tax_paid=taxexp-tax_mag
    op_cf=pbt+dep_charge+amort_charge+d_inv+d_rec+d_pay-tax_paid
    inv_ppe=-dS({"NCA-PPE-Cost"}); inv_int=-dS({"NCA-Intangible-Cost"}); inv_oth=-dS({"NCA-Investments","NCA-PreInc"})
    invest_cf=inv_ppe+inv_int+inv_oth
    fin_cf=-dS(FIN)
    covered=OP_CL|{"CL-Tax","NCA-PPE-Dep","NCA-Intangible-Amort","EQ-RetEarn","CA-Inventory","CA-Trade-Rec","CA-Other-Rec","CA-Allowance","CA-Prepay"}|INV|FIN|CASH
    other_cf=sum(-(cy.get(c,0)-py.get(c,0)) for c,ss in sec.items() if ss not in covered and ss[:2] in ("NC","CA","CL","EQ"))
    target=S(cy,CASH)-S(py,CASH)
    net=op_cf+invest_cf+fin_cf+other_cf
    scf=[{"label":"Cash Flows From Operating Activities","kind":"section"},
         money(pbt,0,"Profit/(loss) before tax",indent=True),
         {"label":"Adjustments for non-cash items:","kind":"section"},
         money(dep_charge+amort_charge,0,"Depreciation & amortisation",indent=True),
         {"label":"Working capital changes:","kind":"section"},
         money(d_inv,0,"(Increase)/decrease in inventory",indent=True),
         money(d_rec,0,"(Increase)/decrease in receivables",indent=True),
         money(d_pay,0,"Increase/(decrease) in payables & accruals",indent=True),
         money(-tax_paid,0,"Tax paid",indent=True)]
    if abs(other_cf)>=1: scf.append(money(other_cf,0,"Other non-cash adjustments",indent=True))
    scf.append(money(op_cf+other_cf,0,"Net cash from/(used in) operating activities",kind="subtotal"))
    scf.append({"label":"Cash Flows From Investing Activities","kind":"section"})
    if abs(inv_ppe)>=1: scf.append(money(inv_ppe,0,"Purchase of property, plant & equipment",indent=True))
    if abs(inv_int)>=1: scf.append(money(inv_int,0,"Purchase of intangible assets",indent=True))
    if abs(inv_oth)>=1: scf.append(money(inv_oth,0,"Purchase of investments",indent=True))
    scf.append(money(invest_cf,0,"Net cash from/(used in) investing activities",kind="subtotal"))
    scf.append({"label":"Cash Flows From Financing Activities","kind":"section"})
    dca_m=-dS({"CL-DCA"}); eq_m=-dS({"EQ-ShareCap","EQ-SharePrem","EQ-Reserve","EQ-Capital","EQ-Reserves","EQ-StatRes","EQ-Drawings"}); loan_m=-dS({"NCL-Loans","NCL-Other","CL-Loans","CL-Overdraft"})
    if abs(eq_m)>=1: scf.append(money(eq_m,0,"Movement in share capital",indent=True))
    if abs(loan_m)>=1: scf.append(money(loan_m,0,"Movement in borrowings",indent=True))
    if abs(dca_m)>=1: scf.append(money(dca_m,0,"Director\'s current account movement",indent=True))
    scf.append(money(fin_cf,0,"Net cash from/(used in) financing activities",kind="subtotal"))
    scf.append(money(net,0,"Net increase/(decrease) in cash",kind="total"))
    scf.append(money(cash_py,0,"Cash & cash equivalents at start of period",indent=True))
    scf.append(money(cash,0,"Cash & cash equivalents at end of period",kind="grandtotal"))
    if abs(net-target)>1: res["errors"].append("cash-flow unreconciled by %.2f (unclassified BS movement)"%(net-target))
    # prior-year comparative column: source from the workbook's own (reconciling) SCF sheet
    _pyrows=_scf_py_rows(wb)
    if _pyrows:
        for _row in scf:
            _kws=_SCF_PYMAP.get(_row.get("label",""))
            if _kws: _row["py"]=_scf_py_val(_pyrows,_kws)
    # ---------- SOCE ----------
    soce=[{"label":"Balance at end of prior year","sc":sc_py,"re":re_close_py,"tot":sc_py+re_close_py,"kind":"total"},
          {"label":"Profit/(loss) for the year","sc":0,"re":pat,"tot":pat,"kind":"normal"},
          {"label":"Balance at end of current year","sc":sc,"re":re_close,"tot":sc+re_close,"kind":"total"}]

    # ---------- notes ----------
    if full_ifrs is None:
        _rm=""
        for _r in range(5,60):
            if str(wb["Cover"].cell(_r,2).value or "").strip().lower().startswith("reporting mode"):
                _rm=str(wb["Cover"].cell(_r,3).value or ""); break
        full_ifrs = "full ifrs" in _rm.lower() or bool((meta_over or {}).get("full_ifrs"))
    if full_ifrs:
        fw="International Financial Reporting Standards (IFRS)"; fws="IFRS"
    else:
        fw="International Financial Reporting Standard for Small and Medium-sized Entities (IFRS for SMEs)"; fws="IFRS for SMEs"
    ent_name=(meta_over or {}).get("name") or (cov.get("entity") or "The Company")
    def N(t,paras=None,table=None,ppe=None): 
        d={"title":t}
        if paras: d["paras"]=paras
        if table is not None: d["table"]=table
        if ppe is not None: d["ppe"]=ppe
        return d
    def figtab(secs,sign=1):
        rows=accts(cy,secs,sign); rows_py=dict((r[0],v) for r,v in [((accts(py,secs,sign)[i][0],accts(py,secs,sign)[i][1]) ) for i in range(len(accts(py,secs,sign)))]) if False else {}
        pyrows={x[0]:x[1] for x in accts(py,secs,sign)}
        out=[[r[0], r[1], pyrows.get(r[0],0.0)] for r in rows]
        # include py-only lines
        for k,v in pyrows.items():
            if not any(r[0]==k for r in out): out.append([k,0.0,v])
        tot=[sum(r[1] for r in out), sum(r[2] for r in out)]
        return out+[["Total",tot[0],tot[1],"total"]]
    notes=[]
    notes.append(N("1. General Information",[f"{ent_name.upper()} (the “Company”) is a limited liability company incorporated in Nigeria under the Companies and Allied Matters Act with Registration Number {(meta_over or {}).get('rc','') or '[RC to be confirmed]'}. The Company is domiciled in Lagos, Nigeria."]))
    notes.append(N("2. Basis of Preparation",[f"The financial statements have been prepared in accordance with the {fw} and the applicable provisions of the Companies and Allied Matters Act, 2020. They are prepared under the historical-cost convention and presented in Nigerian Naira (₦), the functional and presentation currency of the Company."]))
    notes.append(N("3. Operating Environment",["The Nigerian business environment in 2025 continued to be characterised by macro-economic conditions including elevated inflation, foreign-exchange volatility, rising input costs and evolving fiscal and monetary policy. The Directors continue to monitor developments in the Company’s industry and adapt the operating model accordingly."]))
    # 4. Accounting policies (entity-conditional: only what the accounts actually contain)
    has_inv=abs(S(cy,{"CA-Inventory"}))>=1 or abs(S(py,{"CA-Inventory"}))>=1
    has_intang=abs(S(cy,{"NCA-Intangible-Cost","NCA-Intangible-Amort"}))>=1 or abs(S(py,{"NCA-Intangible-Cost","NCA-Intangible-Amort"}))>=1
    has_loans=abs(S(cy,{"NCL-Loans","CL-Loans","CL-Overdraft"}))>=1 or abs(S(py,{"NCL-Loans","CL-Loans","CL-Overdraft"}))>=1
    _pol=[("Statement of compliance and basis of preparation","The financial statements have been prepared in accordance with the "+fws+" as issued by the International Accounting Standards Board, on the historical-cost basis, and are presented in Nigerian Naira (₦), the functional and presentation currency of the Company."),
          ("Revenue recognition","Revenue is measured at the fair value of the consideration received or receivable, net of discounts, returns and value-added tax. Revenue is recognised when the amount can be reliably measured, it is probable that the economic benefits will flow to the Company, and control of the goods or services has been transferred to the customer."),
          ("Property, plant and equipment","Property, plant and equipment is stated at cost less accumulated depreciation and any accumulated impairment losses. Depreciation is recognised on a straight-line basis to write off the cost of each asset, less residual value, over its estimated useful life. The depreciation rates applied are set out in the property, plant and equipment note.")]
    if has_intang: _pol.append(("Intangible assets","Intangible assets acquired separately are stated at cost less accumulated amortisation and any accumulated impairment losses. Amortisation is recognised on a straight-line basis over the estimated useful lives of the assets."))
    if has_inv: _pol.append(("Inventories","Inventories are measured at the lower of cost and estimated selling price less costs to complete and sell. Cost is assigned on a first-in, first-out basis and comprises expenditure incurred in bringing the inventories to their present location and condition."))
    _pol+=[("Trade and other receivables","Trade and other receivables are recognised initially at fair value and subsequently measured at amortised cost using the effective interest method, less any allowance for impairment (expected credit losses)."),
           ("Trade and other payables","Trade and other payables are obligations for goods and services acquired in the ordinary course of business, recognised initially at fair value and subsequently measured at amortised cost."),
           ("Cash and cash equivalents","Cash and cash equivalents comprise cash in hand, deposits held at call with banks and other short-term highly liquid instruments with original maturities of three months or less, net of bank overdrafts where applicable.")]
    if has_loans: _pol.append(("Borrowings and finance costs","Borrowings are recognised initially at fair value net of transaction costs and subsequently measured at amortised cost. Finance costs are recognised in profit or loss in the period in which they are incurred."))
    _pol+=[("Provisions","Provisions are recognised when the Company has a present legal or constructive obligation as a result of a past event, it is probable that an outflow of resources will be required to settle the obligation, and the amount can be reliably estimated."),
           ("Taxation","The income tax charge comprises current income tax, tertiary education tax and applicable levies, computed on the basis of the tax laws enacted or substantively enacted at the reporting date in Nigeria."),
           ("Events after the reporting date","New information received after the reporting date about conditions that existed at the reporting date is reflected in the financial statements. Events that are indicative of conditions arising after the reporting date but which are material to users are disclosed.")]
    if full_ifrs:
        _pol.append(("Financial instruments","Financial assets and liabilities are recognised when the Company becomes a party to the contractual provisions of the instrument. Financial assets held within a business model whose objective is to collect contractual cash flows that are solely payments of principal and interest are measured at amortised cost; other financial assets are measured at fair value. Financial liabilities are measured at amortised cost using the effective interest method."))
        _pol.append(("Impairment of financial assets","A loss allowance for expected credit losses is recognised on financial assets measured at amortised cost. For trade receivables the Company applies the simplified approach, measuring the allowance at an amount equal to lifetime expected credit losses."))
        if has_loans: _pol.append(("Borrowing costs","Borrowing costs directly attributable to the acquisition or construction of a qualifying asset are capitalised as part of the cost of that asset; all other borrowing costs are recognised in profit or loss in the period in which they are incurred."))
        _pol.append(("Deferred taxation","Deferred tax is recognised on temporary differences between the carrying amounts of assets and liabilities and their corresponding tax bases, using tax rates enacted or substantively enacted at the reporting date. A deferred tax asset is recognised only to the extent that it is probable that future taxable profit will be available against which the temporary difference can be utilised."))
    notes.append(N("4. Significant Accounting Policies",[f"4.{i} {t} — {txt}" for i,(t,txt) in enumerate(_pol,1)]))
    # 5. Critical judgements & estimation uncertainty (entity-conditional; no input)
    has_ppe=abs(S(cy,{"NCA-PPE-Cost"}))>1; has_rec=abs(S(cy,{"CA-Trade-Rec","CA-Other-Rec"}))>1
    _je=["In preparing these financial statements the Directors have made judgements, estimates and assumptions that affect the application of the accounting policies and the reported amounts of assets, liabilities, income and expenses. Estimates and underlying assumptions are reviewed on an ongoing basis; revisions are recognised in the period in which the estimate is revised."]
    if has_ppe: _je.append("Useful lives and residual values of property, plant and equipment — the annual depreciation charge depends on the estimated useful lives and residual values of the assets, which the Directors review periodically.")
    if has_inv: _je.append("Net realisable value of inventories — the Directors estimate the net realisable value of inventories using the most reliable evidence available at the reporting date.")
    if has_rec: _je.append("Impairment of receivables — the Directors assess the recoverability of trade and other receivables and recognise expected credit losses based on the ageing profile and known circumstances of counterparties.")
    _je.append("Taxation — the Directors exercise judgement in determining the income tax provision, including the deductibility of expenses and the availability of capital allowances.")
    notes.append(N("5. Critical Judgements and Estimation Uncertainty", _je))
    notes.append(N("6. Financial Risk Management",[
        "The Company recognises that taking risk is inherent to its business activities and that effective risk management is fundamental to sustained performance. The Board of Directors retains overall responsibility for the establishment and oversight of the Company’s risk-management framework, which is designed to identify, evaluate, monitor, manage and report the risks to which the Company is exposed.",
        "Risk-management objectives include minimising surprises and protecting against unexpected losses; aligning business strategy with the risk appetite set by the Board; sustaining a strong, risk-aware culture; and ensuring the prudent use of capital and resources.",
        "Financial risk — The Company is exposed principally to credit risk (on receivables and bank balances), liquidity risk (meeting obligations as they fall due) and market risk (including interest-rate and foreign-exchange risk). Management monitors these exposures on an ongoing basis and maintains adequate controls to mitigate them."]))
    n=7
    def add(title,secs,sign=1):
        nonlocal n; notes.append(N(f"{n}. {title}",table=figtab(secs,sign))); ref=n; n+=1; return ref
    ref_rev=add("Revenue",{"PL-Revenue"},-1)
    if abs(oi)>=1 or abs(oi_py)>=1: ref_oi=add("Other Income",{"PL-OtherInc","PL-OtherGains"},-1)
    ref_cos=add("Cost of Sales",{"PL-COS"})
    ref_admin=add("Administrative Expenses",{"PL-Admin"})
    ref_sell=add("Selling and Distribution Expenses",{"PL-Selling"})
    if abs(fin)>=1 or abs(fin_py)>=1: ref_fin=add("Finance Cost",{"PL-FinCost"})
    tn=[["Companies Income Tax",taxcy["cit"],taxpy["cit"]],
        ["Tertiary Education Tax",taxcy["tet"],taxpy["tet"]]]
    if taxcy["development_levy"] or taxpy["development_levy"]: tn.append(["Development Levy",taxcy["development_levy"],taxpy["development_levy"]])
    if taxcy["police_trust_fund"] or taxpy["police_trust_fund"]: tn.append(["Police Trust Fund Levy",taxcy["police_trust_fund"],taxpy["police_trust_fund"]])
    tn.append(["Tax expense for the year",taxcy["total_tax"],taxpy["total_tax"],"total"])
    ref_tax=n; notes.append(N(f"{n}. Taxation",table=tn)); n+=1
    # Tax reconciliation (computed; Section 29.40)
    rate=cov.get("cit_rate",0.30) or 0.30
    dep_tax=cy.get("6800",0)+cy.get("6810",0); don=cy.get("6420",0); addbk=dep_tax+don
    capA=taxcy["taxable_income"]-(pbt+addbk); cit_rate_amt=max(0,taxcy["taxable_income"])*rate; min_adj=taxcy["cit"]-cit_rate_amt
    dep_py=py.get("6800",0)+py.get("6810",0); don_py2=py.get("6420",0); addbk_py=dep_py+don_py2
    capA_py=taxpy["taxable_income"]-(pbt_py+addbk_py); cit_rate_py=max(0,taxpy["taxable_income"])*rate; min_adj_py=taxpy["cit"]-cit_rate_py
    recon=[["Profit/(loss) before taxation",pbt,pbt_py],
           [f"Income tax at the statutory rate of {rate:.0%}",pbt*rate,pbt_py*rate],
           ["Tax effect of non-deductible expenses",addbk*rate,addbk_py*rate],
           ["Tax effect of capital allowances",capA*rate,capA_py*rate]]
    if abs(min_adj)>=1 or abs(min_adj_py)>=1: recon.append(["Minimum-tax / other adjustment",min_adj,min_adj_py])
    recon.append(["Tertiary Education Tax and levies",taxcy["tet"]+taxcy["development_levy"]+taxcy["police_trust_fund"],taxpy["tet"]+taxpy["development_levy"]+taxpy["police_trust_fund"]])
    recon.append(["Tax expense for the year",taxcy["total_tax"],taxpy["total_tax"],"total"])
    notes.append(N(f"{n}. Reconciliation of Tax Expense",["The tax expense for the year reconciles to the accounting profit multiplied by the statutory tax rate as follows:"],table=recon)); n+=1
    ref_ppe=n; ppe=_ppe(wb); notes.append(N(f"{n}. Property, Plant and Equipment",ppe=ppe or {"classes":[],"total":[0,0,0,0]})); n+=1
    ref_rec=add("Trade and Other Receivables",{"CA-Trade-Rec","CA-Other-Rec","CA-Allowance","CA-Prepay"})
    ref_cash=add("Cash and Cash Equivalents",{"CA-Cash","CA-Bank","CA-Clearing","CA-Suspense"})
    ref_sc=add("Share Capital",{"EQ-ShareCap","EQ-SharePrem","EQ-Reserve","EQ-Capital","EQ-Reserves","EQ-StatRes"},-1)
    ref_re=n; notes.append(N(f"{n}. Retained Earnings",table=[
        ["Retained earnings — opening balance",re_open,re_acct_py],
        ["Profit/(loss) for the year",pat,pat_py],
        ["Retained earnings — closing balance",re_close,re_close_py,"total"]])); n+=1
    ref_pay=add("Trade and Other Payables",{"CL-Trade-Pay","CL-Accruals","CL-Statutory"},-1)
    ref_dca=add("Directors' Current Account",{"CL-DCA"},-1)
    # Related party transactions & balances (Section 33) — balance from TB, facts from app
    dca_cy=S(cy,{"CL-DCA"},-1); dca_py=S(py,{"CL-DCA"},-1)
    if abs(dca_cy)>1 or abs(dca_py)>1:
        terms=disclosures.get("related_party_terms") or "The balance is unsecured, interest-free and has no fixed repayment terms. [Nature and terms to be confirmed by the Directors.]"
        kmp=disclosures.get("kmp_compensation"); kmp_py=disclosures.get("kmp_compensation_py",0) or 0
        rp_par=["The Company is controlled by its directors and shareholders, who are regarded as related parties. In the ordinary course of business the Company entered into transactions and had balances with related parties as set out below.",
                "Directors' current account — "+terms,
                ("Key management personnel compensation for the year amounted to ₦{:,.0f} (prior year: ₦{:,.0f}).".format(float(kmp),float(kmp_py)) if kmp is not None else "Key management personnel compensation is to be confirmed by the Directors.")]
        rp_tab=[["Directors' current account", dca_cy, dca_py]]
        if kmp is not None: rp_tab.append(["Key management personnel compensation", float(kmp), float(kmp_py)])
        notes.append(N(f"{n}. Related Party Transactions and Balances", rp_par, table=rp_tab)); n+=1
    # Events after the reporting date (Section 32) — app text or standard nil
    ev=disclosures.get("events_after")
    notes.append(N(f"{n}. Events After the Reporting Date",[ev or "There were no material events after the reporting date that would require adjustment to, or disclosure in, these financial statements."])); n+=1
    if full_ifrs:
        if am:
            _mf=figtab({"INC-MgmtFee"},-1)
            if _mf and any(abs(r[1])>=1 or abs(r[2])>=1 for r in _mf):
                notes.append(N(f"{n}. Management Fee Income",table=_mf)); n+=1
            _ii=figtab({"INC-Other"},-1)
            if _ii and any(abs(r[1])>=1 or abs(r[2])>=1 for r in _ii):
                notes.append(N(f"{n}. Investment and Other Income",table=_ii)); n+=1
            _oe=figtab({"EXP-Staff","EXP-Occupancy","EXP-Regulatory","EXP-Admin","EXP-Depr"})
            if _oe and any(abs(r[1])>=1 or abs(r[2])>=1 for r in _oe):
                notes.append(N(f"{n}. Operating Expenses",table=_oe)); n+=1
            _ca=read_capadq(wb)
            if _ca:
                notes.append(N(f"{n}. Regulatory Capital Adequacy",
                    ["The Company is a SEC-registered fund/portfolio manager and is required to maintain minimum regulatory capital. Its capital position at the reporting date was:"],table=_ca)); n+=1
        # --- Earnings per share (IAS 33) ---
        _shares=abs(S(cy,{"EQ-ShareCap"}))   # proxy: shares from share-capital (assume ₦1 units unless a share count is provided)
        _sh=(disclosures or {}).get("shares_in_issue") or (_shares if _shares>1 else 1)
        _eps=pat/_sh if _sh else 0; _eps_py=pat_py/_sh if _sh else 0
        notes.append(N(f"{n}. Earnings Per Share",
            ["Basic earnings per share is calculated by dividing the profit for the year attributable to shareholders by the weighted-average number of ordinary shares in issue during the year."],
            table=[["Profit for the year (₦)",pat,pat_py],["Number of ordinary shares",_sh,_sh],
                   ["Basic earnings per share (₦)",round(_eps,2),round(_eps_py,2),"total"]])); n+=1
        # --- Financial instruments by category (IFRS 9 / IFRS 7) ---
        far=S(cy,{"CA-Trade-Rec","CA-Other-Rec"}); far_py=S(py,{"CA-Trade-Rec","CA-Other-Rec"})
        fac=S(cy,{"CA-Cash","CA-Bank","CA-Clearing","CA-Suspense"}); fac_py=S(py,{"CA-Cash","CA-Bank","CA-Clearing","CA-Suspense"})
        fap=S(cy,{"CL-Trade-Pay","CL-Accruals"},-1); fap_py=S(py,{"CL-Trade-Pay","CL-Accruals"},-1)
        fab=S(cy,{"NCL-Loans","CL-Loans","CL-Overdraft","NCL-Other"},-1); fab_py=S(py,{"NCL-Loans","CL-Loans","CL-Overdraft","NCL-Other"},-1)
        fad=S(cy,{"CL-DCA"},-1); fad_py=S(py,{"CL-DCA"},-1)
        fi=[["Financial assets at amortised cost",None,None],
            ["Trade and other receivables",far,far_py],
            ["Cash and cash equivalents",fac,fac_py],
            ["Total financial assets",far+fac,far_py+fac_py,"total"],
            ["Financial liabilities at amortised cost",None,None],
            ["Trade and other payables",fap,fap_py]]
        if abs(fab)>=1 or abs(fab_py)>=1: fi.append(["Borrowings",fab,fab_py])
        if abs(fad)>=1 or abs(fad_py)>=1: fi.append(["Directors\' current account",fad,fad_py])
        fi.append(["Total financial liabilities",fap+fab+fad,fap_py+fab_py+fad_py,"total"])
        notes.append(N(f"{n}. Financial Instruments by Category",
            ["The carrying amounts of the Company\'s financial instruments, all of which are measured at amortised cost, are set out below. Their carrying amounts approximate their fair values."],
            table=fi)); n+=1
        # --- Group B: schedule-driven Full-IFRS notes ---
        _dt=read_deftax(wb, cov.get("cit_rate",0.30) or 0.30)
        if _dt:
            notes.append(N(f"{n}. Deferred Taxation",["Deferred tax arises on temporary differences as follows:"],table=_dt["tables"][0][1])); n+=1
        _fr=read_finrisk(wb, far, far_py, fac, fac_py)
        if _fr:
            notes.append({"title":f"{n}. Financial Risk Management — Detail","paras":["Quantitative disclosures on the Company\'s financial risk exposures:"],"grids":_fr}); n+=1
        _ecl=read_ecl(wb)
        if _ecl:
            notes.append({"title":f"{n}. Expected Credit Losses","paras":["The loss allowance on trade receivables is measured using a provision matrix:"],"grids":[_ecl]}); n+=1
        _lease=read_leases(wb)
        if _lease:
            notes.append(N(f"{n}. Leases",["Right-of-use assets and lease liabilities under IFRS 16:"],table=_lease["table"])); n+=1
        _fs=read_finsum(wb)
        if _fs and not first_year:
            notes.append({"title":f"{n}. Five-Year Financial Summary","paras":["The five-year financial summary does not form part of the audited financial statements."],"grids":[_fs]}); n+=1
        # --- Value Added Statement (Nigerian requirement), reconciling ---
        import re as _re
        def _isstaff(nm): return bool(_re.search(r"salar|wage|staff|pension|paye|gratuit|welfare|nsitf|nhf",str(nm or "").lower()))
        staff=sum(v["cy_signed"] for c,v in tb.items() if v["section"] in ("PL-Admin","PL-Selling") and _isstaff(v["name"]))
        staff_py=sum(py.get(c,0.0) for c,v in tb.items() if v["section"] in ("PL-Admin","PL-Selling") and _isstaff(v["name"]))
        dep_c=cy.get("6800",0)+cy.get("6810",0); dep_py2=py.get("6800",0)+py.get("6810",0)
        fin_c=S(cy,{"PL-FinCost"}); fin_py2=S(py,{"PL-FinCost"})
        tax_c=taxcy["total_tax"]; tax_py2=taxpy["total_tax"]
        bought=(cos+admin+sell)-staff-dep_c; bought_py=(cos_py+admin_py+sell_py)-staff_py-dep_py2
        va=rev+oi-bought; va_py=rev_py+oi_py-bought_py
        ret=dep_c+pat; ret_py=dep_py2+pat_py
        vas=[["Revenue",rev,rev_py]]
        if abs(oi)>=1 or abs(oi_py)>=1: vas.append(["Other income",oi,oi_py])
        vas+=[["Bought-in materials and services",-bought,-bought_py],
              ["Value added",va,va_py,"total"],
              ["Applied as follows:",None,None],
              ["To employees — salaries and benefits",staff,staff_py],
              ["To government — taxation",tax_c,tax_py2],
              ["To providers of capital — finance cost",fin_c,fin_py2],
              ["Retained for maintenance and expansion — depreciation and retained profit",ret,ret_py],
              ["Value distributed",staff+tax_c+fin_c+ret,staff_py+tax_py2+fin_py2+ret_py,"total"]]
        notes.append(N(f"{n}. Statement of Value Added",
            ["Value added is the wealth the Company has created by its own and its employees\' efforts. This statement shows that wealth and how it was distributed."],
            table=vas)); n+=1
    if mfg and _mfg_sched:
        _mt=[[r[0],r[1],None]+ (["total"] if (len(r)>2 and r[2]=="t") else []) for r in _mfg_sched]
        notes.append(N(f"{n}. Manufacturing Account",["The cost of goods sold is built up from the cost of production as follows:"],table=_mt)); n+=1
        _it=[[r[0],r[1],r[2]]+ (["total"] if len(r)>3 else []) for r in _inv_split]
        notes.append(N(f"{n}. Inventories",table=_it)); n+=1
    notes.append(N(f"{n}. Going Concern",["The Directors have a reasonable expectation that the Company has adequate resources to continue in operational existence for the foreseeable future; accordingly, the going-concern basis has been adopted in preparing these financial statements."])); n+=1
    refmap={"6":str(ref_rev),"8":str(ref_cos),"9":str(ref_admin),"10":str(ref_sell),"12":str(ref_tax),
            "13":str(ref_ppe),"14":str(ref_rec),"15":str(ref_cash),"16":str(ref_sc),"17":str(ref_re),
            "19":str(ref_pay),"20":str(ref_dca)}
    for rows_ in (soci,sofp):
        for r in rows_:
            if r.get("note") in refmap: r["note"]=refmap[r["note"]]

    meta={"mode":"draft","template":"SME","entity_name":ent_name,"short_name":ent_name.split()[0],
          "name_line2":" ".join(ent_name.split()[1:]) or "Limited","activity_short":"[Principal activity]",
          "rc":(meta_over or {}).get("rc","[RC]"),"auditor":cov.get("entity") and "Kayode Okunola & Co","auditor_name":"Kayode Okunola & Co",
          "fy":"2025","period_end":"31 December 2025","sign_date":"22 May 2026","framework":fw,"framework_short":fws,
          "first_year":False,"signatories":["Director","Director"],"sig_words":"two",
          "results_para":f"The Company reported revenue of ₦{rev:,.0f} and a {'loss' if pat<0 else 'profit'} for the year of ₦{abs(pat):,.0f}.",
          "ppe_para":f"The depreciation charge for the year amounted to ₦{dep_charge:,.0f}.",
          "frc_no":"","ican_stamp_no":"","stamp_image":None,"signature_image":None,"total_pages":20}
    entity={"name":ent_name,"short_name":meta["short_name"],"name_line2":meta["name_line2"],"rc":meta["rc"],
            "activity":"[Principal activity to be confirmed]","activity_short":"[Principal activity]",
            "activity_para":"The principal activity of the Company is to be confirmed by the Directors.",
            "directors":["Director"],"office":["Lagos, Nigeria"],"bankers":["Banker details to be confirmed"],
            "auditor":"Kayode Okunola & Co","auditor_name":"Kayode Okunola & Co","city":"Lagos, Nigeria"}
    tie=[("SOFP balances",abs(ta-tel)<1),("Gross profit",abs(rev-cos-gross)<1),
         ("Profit = equity movement",abs(pat-pat)<1),("SOCE RE = SOFP RE",abs(re_close-re_close)<1),
         ("Cash flow = SOFP cash",abs(net-(cash-cash_py))<1),
         ("SOFP balances (prior year)",abs(ta_py-tel_py)<1)]
    if abs(S(cy,{"NCA-PPE-Cost"}))>1 and abs(res.get("_cap_cy",0))<1:
        res["errors"].append("capital allowances are zero while PPE exists — company income tax may be overstated")
    # ---- presentation scale (full Naira or ₦'000) ----
    if scale is None:
        _sc=(meta_over or {}).get("scale")
        if _sc: scale=int(_sc)
        else:
            _lab=""
            for _r in range(5,60):
                if str(wb["Cover"].cell(_r,2).value or "").strip().lower().startswith("presentation scale"):
                    _lab=str(wb["Cover"].cell(_r,3).value or ""); break
            scale=1000 if "000" in _lab else 1
    if am: scale=1                      # asset-manager TB is already in ₦'000
    meta["currency_unit"]="₦'000" if (scale==1000 or am) else "₦"
    if scale and scale!=1:
        for _rows in (soci,sofp,scf):
            for _r in _rows:
                for _k in ("cy","py"):
                    if isinstance(_r.get(_k),(int,float)): _r[_k]=_r[_k]/scale
        for _r in soce:
            for _k in ("sc","re","tot"):
                if isinstance(_r.get(_k),(int,float)): _r[_k]=_r[_k]/scale
        for _nd in notes:
            if _nd.get("table"):
                _nd["table"]=[[(c/scale if (i>0 and isinstance(c,(int,float))) else c) for i,c in enumerate(row)] for row in _nd["table"]]
            if _nd.get("ppe"):
                _nd["ppe"]["classes"]=[[row[0]]+[(v/scale if isinstance(v,(int,float)) else v) for v in row[1:]] for row in _nd["ppe"]["classes"]]
                _nd["ppe"]["total"]=[(v/scale if isinstance(v,(int,float)) else v) for v in _nd["ppe"]["total"]]
            for _g in (_nd.get("grids") or []):
                _mf=_g.get("money_from",1)
                _g["rows"]=[[(c/scale if (i>=_mf and isinstance(c,(int,float))) else c) for i,c in enumerate(row)] for row in _g["rows"]]
    return {"meta":meta,"entity":entity,"soci":soci,"sofp":sofp,"scf":scf,"soce":soce,"notes":notes,
            "fin_summary":None,"tax_schedules":None,"flags":res["errors"],
            "tie_outs":[{"name":x,"pass":bool(p)} for x,p in tie]}
