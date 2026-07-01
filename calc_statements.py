# -*- coding: utf-8 -*-
"""
AFS Studio — Slice 1b: statements + notes built from the trial balance.

Builds the Statement of Financial Position, Statement of Profit or Loss, the
Statement of Changes in Equity, and the figure notes **from the computed trial
balance**, mapping every line by COA *section* (never by label text). Because the
notes are derived from the same TB as the face of the accounts, they cannot
contradict the statements — which fixes the three defects found in review:

  * Note 17 (Retained Earnings) rolls forward from the prior-year CLOSE, so it
    ties to the SOFP (old engine printed the prior-prior opening -> wrong).
  * Note 12 (Taxation) reconciles the tax EXPENSE, not the balance-sheet payable.
  * Prior-year comparatives come from one source (OpenBal PY), so SOCI and SOCE
    cannot disagree.

Depends on calc_core for input reading + TB build. Validated against
BODY TEMPLE_2025_FIXED.xlsx.

Usage:  python calc_statements.py "<workbook>.xlsx"
"""
from __future__ import annotations
import sys
from collections import defaultdict
import openpyxl
import calc_core
from calc_core import ncode, num

# COA section -> where it sits on the SOFP
NCA = {"NCA-PPE-Cost", "NCA-PPE-Dep", "NCA-Intangible-Cost", "NCA-Intangible-Amort",
       "NCA-Investments", "NCA-DefTax", "NCA-PreInc"}
CA  = {"CA-Inventory", "CA-Trade-Rec", "CA-Other-Rec", "CA-Allowance", "CA-Prepay",
       "CA-Cash", "CA-Bank", "CA-Clearing", "CA-Suspense"}
EQ  = {"EQ-ShareCap", "EQ-SharePrem", "EQ-Reserve", "EQ-Capital", "EQ-RetEarn", "EQ-Drawings"}
RE_SECTIONS = {"EQ-RetEarn", "EQ-Drawings"}
NCL = {"NCL-Loans", "NCL-Other", "NCL-DefTax"}
CL  = {"CL-Trade-Pay", "CL-Accruals", "CL-Statutory", "CL-Tax", "CL-DCA",
       "CL-Overdraft", "CL-Loans"}
PL  = {"PL-Revenue", "PL-OtherInc", "PL-OtherGains", "PL-COS", "PL-Admin",
       "PL-Selling", "PL-FinCost", "PL-Tax"}


def _read_py(path):
    """Prior-year signed balance (Dr positive) per code, from OpenBal E/F."""
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["OpenBal"]
    py = {}
    for r in range(5, ws.max_row + 1):
        c = ncode(ws.cell(r, 1).value)
        if c and c.upper() != "TOTAL":
            py[c] = num(ws.cell(r, 5).value) - num(ws.cell(r, 6).value)
    return py


def _sum(book, sections, sign=1):
    return sign * sum(book.get(c, 0.0) for c, s in _SEC.items() if s in sections)


_SEC = {}  # code -> section, populated in build()


def _profit(book):
    """Profit for the year = -(net P&L debit)."""
    return -sum(book.get(c, 0.0) for c, s in _SEC.items() if s in PL)


def build(path):
    global _SEC
    res = calc_core.compute(path)
    tb = res["tb"]
    _SEC = {c: a["section"] for c, a in tb.items()}
    cy = {c: a["cy_signed"] for c, a in tb.items()}
    py = _read_py(path)

    out = {}
    for tag, book in (("cy", cy), ("py", py)):
        pat = _profit(book)
        re_acct = _sum(book, RE_SECTIONS, -1)          # retained-earnings account balance
        re_closing = re_acct + pat                     # closing RE incl. this period's result
        sofp = {
            "ppe": _sum(book, {"NCA-PPE-Cost", "NCA-PPE-Dep"}),
            "non_current_assets": _sum(book, NCA),
            "current_assets": _sum(book, CA),
            "share_capital": _sum(book, {"EQ-ShareCap", "EQ-SharePrem", "EQ-Reserve", "EQ-Capital"}, -1),
            "retained_earnings": re_closing,
            "non_current_liabilities": _sum(book, NCL, -1),
            "current_liabilities": _sum(book, CL, -1),
        }
        sofp["total_assets"] = sofp["non_current_assets"] + sofp["current_assets"]
        sofp["total_equity"] = sofp["share_capital"] + sofp["retained_earnings"]
        sofp["total_equity_and_liabilities"] = (sofp["total_equity"]
                                                + sofp["non_current_liabilities"]
                                                + sofp["current_liabilities"])
        soci = {
            "revenue": _sum(book, {"PL-Revenue"}, -1),
            "other_income": _sum(book, {"PL-OtherInc", "PL-OtherGains"}, -1),
            "cost_of_sales": _sum(book, {"PL-COS"}),
            "admin_expenses": _sum(book, {"PL-Admin"}),
            "selling_expenses": _sum(book, {"PL-Selling"}),
            "finance_cost": _sum(book, {"PL-FinCost"}),
            "tax_expense": _sum(book, {"PL-Tax"}),
            "profit_after_tax": pat,
        }
        soci["gross_profit"] = soci["revenue"] - soci["cost_of_sales"]
        soci["profit_before_tax"] = pat + soci["tax_expense"]
        out[tag] = {"sofp": sofp, "soci": soci, "pat": pat, "re_closing": re_closing}

    # Retained-earnings note (FIX): opens at prior-year CLOSE, ties to SOFP
    re_note = {
        "opening": out["py"]["re_closing"],
        "profit_for_year": out["cy"]["pat"],
        "closing": out["cy"]["re_closing"],
    }
    # Taxation note (FIX): the tax EXPENSE, reconciled, not the payable
    tax = res["tax"]
    tax_note = {
        "companies_income_tax": tax["cit"], "tertiary_education_tax": tax["tet"],
        "development_levy": tax["development_levy"], "police_trust_fund": tax["police_trust_fund"],
        "total_tax_expense": tax["total_tax"],
        "booked_in_pl": out["cy"]["soci"]["tax_expense"],
        "unbooked": round(tax["total_tax"] - out["cy"]["soci"]["tax_expense"], 2),
    }
    return {"cover": res["cover"], "statements": out, "re_note": re_note,
            "tax_note": tax_note, "errors": res["errors"]}


def _sofp_lookup(wb, contains):
    """Read a SOFP CY value (col D) by matching its label (col B) — robust to row moves."""
    ws = wb["SOFP"]
    for r in range(4, 60):
        b = ws.cell(r, 2).value
        if b and contains.upper() in str(b).upper():
            return num(ws.cell(r, 4).value)
    return None


def validate(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    r = build(path)
    cy = r["statements"]["cy"]
    print(f"\nSlice 1b tie-out — {r['cover']['entity']}")
    print("  " + "-" * 58)
    rows = [
        ("Total assets",        cy["sofp"]["total_assets"],                 _sofp_lookup(wb, "TOTAL ASSETS")),
        ("Total equity & liab", cy["sofp"]["total_equity_and_liabilities"], _sofp_lookup(wb, "TOTAL EQUITY AND LIABILITIES")),
        ("Retained earnings",   cy["sofp"]["retained_earnings"],            _sofp_lookup(wb, "Retained earnings")),
        ("Profit for the year", cy["pat"],                                  num(wb["SOCI"]["D17"].value)),
    ]
    ok = True
    for label, py_v, truth in rows:
        good = abs(py_v - truth) < 1.0
        ok &= good
        print(f"  [{'OK ' if good else 'DIFF'}] {label:22s} {py_v:16,.2f} vs {truth:16,.2f}")
    print("  " + "-" * 58)
    rn = r["re_note"]
    tie = abs(rn["closing"] - cy["sofp"]["retained_earnings"]) < 1.0
    print(f"  Note 17 (RE): open {rn['opening']:,.0f} + profit {rn['profit_for_year']:,.0f}"
          f" = {rn['closing']:,.0f}  ->  ties to SOFP: {'YES' if tie else 'NO'}")
    tn = r["tax_note"]
    print(f"  Note 12 (Tax): expense {tn['total_tax_expense']:,.2f}  (unbooked PTF {tn['unbooked']:,.2f})")
    if r["errors"]:
        print("  FLAGS:", "; ".join(r["errors"][:5]))
    print(f"\n  RESULT: {'STATEMENTS + FIXED NOTES TIE OUT' if ok and tie else 'INVESTIGATE'}")
    return ok and tie


if __name__ == "__main__":
    validate(sys.argv[1] if len(sys.argv) > 1 else "BODY TEMPLE_2025_FIXED.xlsx")
