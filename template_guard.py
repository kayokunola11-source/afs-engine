# -*- coding: utf-8 -*-
"""Template / framework guard. Reads the uploaded workbook's Cover to determine its
framework (from the machine-readable Template tag, else the Reporting mode) and its
variant/version, and compares to what the engagement expected. Rejects a clear
mismatch or a non-template file BEFORE anything is generated."""
import openpyxl

FW_NAME={"sme":"IFRS for SMEs","full_ifrs":"Full IFRS"}
CORE_SHEETS={"Cover","COA","TB","SOCI","SOFP"}

def _cover(wb, label_contains):
    ws=wb["Cover"]
    for row in ws.iter_rows(min_row=1,max_row=90,max_col=3,values_only=True):
        b=row[1]; v=row[2]
        if b and label_contains.lower() in str(b).strip().lower() and v not in (None,""):
            return v
    return None

def _fw_from_mode(mode):
    if not mode: return None
    m=str(mode).lower()
    if "full ifrs" in m: return "full_ifrs"
    if "sme" in m or "small and medium" in m: return "sme"
    if "business name" in m or "sole prop" in m: return "sme"   # PIT mode is an SME-family template
    return None

def detect(wb):
    tag=_cover(wb,"template tag")
    fw=var=ver=None
    if tag and "|" in str(tag):
        p=[x.strip().lower() for x in str(tag).split("|")]
        fw=p[0] or None; var=p[1] if len(p)>1 else None; ver=p[2] if len(p)>2 else None
    mode_fw=_fw_from_mode(_cover(wb,"reporting mode") or _cover(wb,"reporting framework"))
    return {"tag_framework":fw,"variant":var,"version":(ver or "").lstrip("v") or None,"mode_framework":mode_fw,
            "framework":fw or mode_fw}

def check(path, expected_framework=None, expected_variant=None, expected_version=None):
    try:
        wb=openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        return {"ok":False,"reject":True,"reason":"The uploaded file could not be read as an Excel workbook (%s)."%e}
    missing=CORE_SHEETS-set(wb.sheetnames)
    if missing:
        return {"ok":False,"reject":True,
                "reason":"This does not look like an AFS template (missing sheet(s): %s). Download the template from the app and prepare your figures on it."%", ".join(sorted(missing))}
    d=detect(wb); actual=d["framework"]; warnings=[]
    ef=(expected_framework or "").strip().lower() or None
    if ef and actual and actual!=ef:
        return {"ok":False,"reject":True,"detected":d,
                "reason":"Framework mismatch — this engagement was set up as %s, but the uploaded workbook is a %s template. Download the %s template and re-prepare before generating."%(
                    FW_NAME.get(ef,ef), FW_NAME.get(actual,actual), FW_NAME.get(ef,ef))}
    if ef and not actual:
        warnings.append("Could not read the workbook's framework (no Template tag or Reporting mode) — proceeded without a framework check.")
    if expected_variant and d["variant"] and d["variant"]!=str(expected_variant).strip().lower():
        warnings.append("Variant differs: engagement expects '%s', workbook says '%s'."%(expected_variant,d["variant"]))
    if expected_version and d["version"] and str(d["version"])!=str(expected_version).strip().lstrip("v"):
        warnings.append("Template version differs: engagement expects v%s, workbook is v%s — consider downloading the current template."%(expected_version,d["version"]))
    return {"ok":True,"reject":False,"detected":d,"warnings":warnings}
