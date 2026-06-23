# -*- coding: utf-8 -*-
"""JUKES-dialect reader: workbooks with lowercase statement tabs
(sopl / sofp / cashflow / socie / notes), label in column A, no Entity/Cover sheet.
Entity details come from the app (entity_overrides)."""
import openpyxl, datetime, re

def _num(v):
    if isinstance(v,(int,float)): return float(v)
    return None

def detect_dialect(xlsx_path):
    wb=openpyxl.load_workbook(xlsx_path, data_only=True, read_only=True)
    sl={s.strip().lower() for s in wb.sheetnames}; wb.close()
    if "sopl" in sl and "sofp" in sl: return "jukes"
    if "soci" in sl or "entity" in sl or "cover" in sl: return "template"
    return "template"

def _cy_year(ws):
    for r in ws.iter_rows(min_row=1,max_row=4,max_col=4,values_only=True):
        for c in r:
            s=str(c).strip() if c is not None else ""
            if s.isdigit() and len(s)==4: return s
    return ""

# ---- label normalisation so tie-outs / notes recognise the lines ----
def _norm_pl(lab):
    l=lab.lower().strip()
    if l=="revenue": return "Revenue","normal"
    if l.startswith("cost of sale"): return "Cost of sales","normal"
    if "gross profit" in l: return "GROSS PROFIT","total"
    if l.startswith("other income"): return "Other income","normal"
    if "administrative expenses" in l: return "Administrative expenses","normal"
    if "selling and distribution" in l: return "Selling and distribution expenses","normal"
    if "finance cost" in l: return "Finance cost","normal"
    if "operating profit" in l or l=="profit before tax": return "OPERATING PROFIT/(LOSS) BEFORE TAX","subtotal"
    if "provision for tax" in l or l=="taxation": return "Taxation","normal"
    if "loss for the year" in l or "profit for the year" in l: return "PROFIT/(LOSS) FOR THE YEAR","total"
    if "other comprehensive income" in l: return "Other comprehensive income","normal"
    if "total comprehensive income" in l: return "TOTAL COMPREHENSIVE INCOME FOR THE YEAR","grandtotal"
    return lab.strip(), "normal"

def _norm_bs(lab):
    l=lab.lower().strip()
    secs={"assets":"ASSETS","equity and liabilities":"EQUITY AND LIABILITIES"}
    subsecs={"non current asset","non-current asset","current asset","equity","current liabilities",
             "non current liabilities","non-current liabilities"}
    if l in secs: return secs[l],"section"
    if l in subsecs:
        title={"non current asset":"Non-current assets","non-current asset":"Non-current assets",
               "current asset":"Current assets","equity":"Equity","current liabilities":"Current liabilities",
               "non current liabilities":"Non-current liabilities","non-current liabilities":"Non-current liabilities"}[l]
        return title,"section"
    if l.startswith("total asset"): return "TOTAL ASSETS","grandtotal"
    if "total capital and liabilities" in l or "total equity and liabilities" in l: return "TOTAL EQUITY AND LIABILITIES","grandtotal"
    if l.startswith("total non current") or l.startswith("total non-current"): return "Total non-current assets","subtotal"
    if l.startswith("total current asset"): return "Total current assets","subtotal"
    if l.startswith("total current liabilit"): return "Total current liabilities","subtotal"
    if "capital reserves" in l or l=="total equity": return "Total equity","subtotal"
    if "property, plant" in l or "property,plant" in l: return "Property, plant and equipment","normal"
    if "cash and cash equivalent" in l: return "Cash and cash equivalents","normal"
    if "trade and other receivable" in l: return "Trade and other receivables","normal"
    if "revenue reserve" in l or "retained earning" in l: return "Retained earnings","normal"
    if l=="share capital": return "Share capital","normal"
    if "trade payable" in l: return "Trade and other payables","normal"
    if "provision for tax" in l: return "Current tax payable","normal"
    if "directors' current account" in l or "directors current account" in l or "director's current" in l: return "Directors' current account","normal"
    return lab.strip(),"normal"

def _read_pl(ws):
    rows=[]
    for r in ws.iter_rows(min_row=4,max_row=40,max_col=4,values_only=True):
        lab=r[0]
        if not (isinstance(lab,str) and lab.strip()): continue
        note=r[1]; cy=_num(r[2]); py=_num(r[3])
        name,kind=_norm_pl(lab)
        row={"label":name,"note":(str(note) if note not in (None,"") else ""),"cy":cy or 0,"py":py or 0,"kind":kind}
        if kind=="normal": row["indent"]=False
        rows.append(row)
    return rows

def _read_bs(ws):
    rows=[]
    for r in ws.iter_rows(min_row=4,max_row=45,max_col=4,values_only=True):
        lab=r[0]
        if not (isinstance(lab,str) and lab.strip()): continue
        if lab.strip().lower() in ("₦","note","notes"): continue
        note=r[1]; cy=_num(r[2]); py=_num(r[3])
        name,kind=_norm_bs(lab)
        if kind=="section":
            if rows and rows[-1].get("kind") not in ("section","blank"): rows.append({"kind":"blank"})
            rows.append({"label":name,"kind":"section"}); continue
        if kind=="grandtotal" and rows and rows[-1].get("kind") not in ("blank","section"):
            rows.append({"kind":"blank"})
        row={"label":name,"note":(str(note) if note not in (None,"") else ""),"cy":cy or 0,"py":py or 0,"kind":kind}
        if kind=="normal": row["indent"]=True
        rows.append(row)
    return rows

def _read_cf(ws):
    rows=[]
    SEC=("cashflow from","add back","movement in working capital","items not involving")
    for r in ws.iter_rows(min_row=5,max_row=40,max_col=3,values_only=True):
        lab=r[0]
        if not (isinstance(lab,str) and lab.strip()): continue
        cy=_num(r[1]); py=_num(r[2]); l=lab.lower()
        if any(s in l for s in SEC) and cy is None and py is None:
            if rows and rows[-1].get("kind") not in ("section","blank"): rows.append({"kind":"blank"})
            rows.append({"label":lab.strip().title(),"kind":"section"}); continue
        kind="normal"
        if l.startswith("net cashflow"): kind="subtotal"
        elif "net increase" in l or "net decrease" in l: kind="total"
        elif "as at 31" in l or "end of" in l or l.startswith("cash & cash equivalent as at 31"): kind="grandtotal"
        if kind=="grandtotal" and rows and rows[-1].get("kind") not in ("blank","section"): rows.append({"kind":"blank"})
        row={"label":lab.strip(),"note":"","cy":cy or 0,"py":py or 0,"kind":kind}
        if kind=="normal": row["indent"]=True
        rows.append(row)
    return rows

def _read_soce(ws, pl_loss_cy):
    # columns: A label, B share capital, C retained earnings, D other, E total
    out=[]
    for r in ws.iter_rows(min_row=3,max_row=14,max_col=5,values_only=True):
        lab=r[0]
        if not (isinstance(lab,str) and lab.strip()): continue
        sc=_num(r[1]); re=_num(r[2]); other=_num(r[3]); tot=_num(r[4]); l=lab.lower()
        kind="total" if l.startswith("balance as at december") or l.startswith("balance at 31") else "normal"
        out.append({"label":lab.strip(),"sc":sc or 0,"re":re,"other":other or 0,"tot":tot,"kind":kind})
    # reconstruct any missing (#REF!) retained earnings from total - share capital - other
    for row in out:
        if row["re"] is None:
            if row["tot"] is not None:
                row["re"]=row["tot"] - (row["sc"] or 0) - (row["other"] or 0)
            else:
                row["re"]=0
        if row["tot"] is None: row["tot"]=(row["sc"] or 0)+(row["re"] or 0)+(row["other"] or 0)
    return out


def _kws_for(title):
    t=title.lower()
    m=[("revenue reserve",["retained","reserve"]),("retained",["retained"]),("revenue",["revenue"]),
       ("other income",["other income"]),("cost of sale",["cost of sale"]),("administrative",["administrative"]),
       ("finance",["finance cost"]),("taxation",["tax"]),("receivable",["receivable"]),
       ("cash",["cash and","cash &"]),("share capital",["share capital"]),("payable",["payable"]),
       ("director",["directors"]),("property",["property","plant"]),("fixed asset",["property","plant"]),
       ("long",["long-term","long term"])]
    for k,v in m:
        if k in t: return v
    w=[x for x in t.split() if len(x)>4]
    return w[:1] or [t]

def _parse_jukes_ppe(rows):
    head=rows[0]; names=[]; cols=[]
    for j in range(2,8):
        v=head[j] if j<len(head) else None
        if v and str(v).strip() and "total" not in str(v).lower():
            names.append(str(v).strip()); cols.append(j)
    def findrow(*keys):
        for r in rows:
            b=str(r[1]).strip().lower() if r[1] else ""
            if b and any(k in b for k in keys): return r
        return None
    cost_r=findrow("cost"); charge_r=findrow("charge for the year","charge")
    dep_r=findrow("bal c/f","c/f"); nbv_r=findrow("net book value","nbv")
    def g(r,j): 
        return (_num(r[j]) or 0) if (r and j<len(r)) else 0
    classes=[]
    for nm,j in zip(names,cols):
        cost=g(cost_r,j); charge=g(charge_r,j); dep=g(dep_r,j); nbv=g(nbv_r,j)
        if cost or dep or nbv or charge: classes.append([nm,cost,dep,charge,nbv])
    tot=[sum(c[i] for c in classes) for i in (1,2,3,4)]
    return {"classes":classes,"total":tot}

def build_jukes_notes(wb):
    """Read the JUKES 'notes' sheet into the full detailed note set, numbered sequentially from 6."""
    ns=None
    for sn in wb.sheetnames:
        if sn.strip().lower()=="notes": ns=wb[sn]; break
    if ns is None: return [], []
    rows=[r for r in ns.iter_rows(min_row=1,max_row=200,max_col=8,values_only=True)]
    blocks=[]; cur=None
    for r in rows:
        a=str(r[0]).strip() if r[0] is not None else ""
        b=str(r[1]).strip() if r[1] is not None else ""
        if a and re.match(r'^\d', a):
            if cur: blocks.append(cur)
            cur={"title":b or a,"rows":[r]}
        elif cur is not None:
            cur["rows"].append(r)
    if cur: blocks.append(cur)
    fig=[]; ref=[]; num=6
    for blk in blocks:
        title=re.sub(r'^\d{4}\s+','',blk["title"]).strip()   # drop leading year e.g. "2024 Fixed Asset Schedule"
        is_ppe="fixed asset" in title.lower() or "property" in title.lower()
        if is_ppe:
            ppe=_parse_jukes_ppe(blk["rows"])
            if not (ppe["classes"] or any(ppe["total"])): continue
            fig.append({"title":f"{num}. Property, Plant and Equipment","ppe":ppe}); ref.append((["property","plant","fixed asset"],num)); num+=1
            continue
        items=[]; total=None
        for r in blk["rows"][1:]:
            b=str(r[1]).strip() if r[1] is not None else ""
            cy=_num(r[2]); py=_num(r[3])
            if b:
                if cy is None and py is None: continue
                items.append([b,cy or 0,py or 0])
            elif cy is not None or py is not None:
                total=(cy or 0, py or 0)
        if not items and total is None: continue
        tbl=list(items)
        if total is not None: tbl.append([f"Total {title.lower()}", total[0], total[1], "total"])
        elif items: tbl.append([f"Total {title.lower()}", sum(i[1] for i in items), sum(i[2] for i in items), "total"])
        fig.append({"title":f"{num}. {title}","table":tbl}); ref.append((_kws_for(title),num)); num+=1
    return fig, ref

def get_data_jukes(xlsx_path, mode="draft", first_year=None, n_sig=2, template="SME",
                   auditor="Kayode Okunola & Co (Chartered Accountants)", auditor_name="Kayode Okunola & Co",
                   frc_no="0968263", ican_stamp_no="", stamp_image=None, signature_image=None,
                   entity_overrides=None):
    eo=entity_overrides or {}
    wb=openpyxl.load_workbook(xlsx_path, data_only=True)
    sopl=wb["sopl"] if "sopl" in wb.sheetnames else wb[[s for s in wb.sheetnames if s.lower()=="sopl"][0]]
    def sheet(name):
        for s in wb.sheetnames:
            if s.strip().lower()==name: return wb[s]
        return None
    sopl=sheet("sopl"); sofp=sheet("sofp"); cfs=sheet("cashflow"); soce=sheet("socie")
    fy=_cy_year(sopl) or _cy_year(sofp) or ""
    soci=_read_pl(sopl); sofp_rows=_read_bs(sofp); scf=_read_cf(cfs) if cfs else []
    def gv(rows,lbl,k="cy"):
        for r in rows:
            if r.get("label","").upper()==lbl.upper(): return r.get(k) or 0
        return 0
    pat=gv(soci,"PROFIT/(LOSS) FOR THE YEAR")
    soce_rows=_read_soce(soce, pat) if soce else []

    name=eo.get("name") or "Company"
    period_end=eo.get("period_end") or (f"31 December {fy}" if fy else "")
    rc=eo.get("rc") or ""
    directors=eo.get("directors") or ["Director"]
    activity=eo.get("activity") or "[Principal activity to be confirmed]"
    city=eo.get("city") or "Lagos, Nigeria"
    office=eo.get("office") or [city]
    bankers=eo.get("bankers") or "Banker details to be confirmed"
    sign_date=eo.get("sign_date") or ""
    SIG_WORDS={1:"one",2:"two",3:"three",4:"four"}
    n_sig=max(1,min(n_sig,max(1,len(directors))))
    fw="International Financial Reporting Standard for Small and Medium-sized Entities (IFRS for SMEs)"
    fws="IFRS for SMEs"
    rev_cy=gv(soci,"Revenue"); rev_py=gv(soci,"Revenue","py")
    def naira(v): return "₦{:,.0f}".format(round(v))
    plw=lambda v:"loss" if v<0 else "profit"
    try: pct=abs((rev_cy-rev_py)/rev_py*100); dirn="increase" if rev_cy>=rev_py else "decrease"
    except ZeroDivisionError: pct=0; dirn="change"
    results_para=(f"The operating results of the Company for the year ended {period_end} are set out in the "
                  f"Statement of Profit or Loss and Other Comprehensive Income. The Company reported revenue of "
                  f"{naira(rev_cy)} (prior year: {naira(rev_py)}), representing a {dirn} of approximately {pct:.0f} "
                  f"per cent, and a {plw(gv(soci,'PROFIT/(LOSS) FOR THE YEAR'))} for the year of "
                  f"{naira(abs(gv(soci,'PROFIT/(LOSS) FOR THE YEAR')))}.")

    cos=gv(soci,"Cost of sales","cy"); adm=gv(soci,"Administrative expenses","cy")
    notes=[
        {"title":"1. General Information","paras":[f"{name.upper()} (the “Company”) is a limited liability company incorporated in Nigeria under the Companies and Allied Matters Act"+(f" with Registration Number {rc}" if rc else "")+f". The Company is domiciled in {city}."]},
        {"title":"2. Basis of Preparation","paras":[f"The financial statements have been prepared in accordance with the {fw} and the applicable provisions of the Companies and Allied Matters Act, 2020, under the historical-cost convention, and are presented in Nigerian Naira (₦)."]},
        {"title":"3. Operating Environment","paras":["The Nigerian business environment in "+(fy or "the year")+" continued to be characterised by elevated inflation, foreign-exchange volatility, rising input costs and evolving fiscal and monetary policy. The Directors continue to monitor developments in the Company's sector.","[App note: auto-tailored to the client's industry once principal activity is confirmed.]"]},
        {"title":"4. Significant Accounting Policies","paras":["4.1 Revenue is recognised when control of goods/services transfers to the customer, net of VAT.","4.2 Property, plant and equipment is stated at cost less accumulated depreciation, depreciated on a straight-line basis.","4.3 Trade receivables and payables are measured at amortised cost.","4.4 Taxation is computed under prevailing Nigerian tax legislation."]},
        {"title":"5. Financial Risk Management","paras":["The Company is exposed to financial, operational and market risks; management has procedures to identify, monitor and mitigate them under Board oversight."]},
    ]
    _fig,_refj=build_jukes_notes(wb)
    import afs_notes as _an
    _an.remap_statement_refs(soci,_refj); _an.remap_statement_refs(sofp_rows,_refj)
    _gcn=6+len(_fig)
    notes=notes+_fig+[{"title":f"{_gcn}. Going Concern","paras":["The Directors have assessed the Company's ability to continue as a going concern and have adopted the going-concern basis in preparing these financial statements."]}]
    entity={"name":name,"short_name":name.split()[0] if name else "Company","name_line2":" ".join(name.split()[1:]) or "Limited",
            "rc":rc,"activity":activity,"activity_short":activity if not activity.startswith("[") else "[Principal activity to be confirmed]",
            "activity_para":f"The principal activity of the Company during the year is {activity if not activity.startswith('[') else '[to be confirmed]'}.",
            "directors":directors,"office":office,"bankers":bankers,"auditor":auditor,"auditor_name":auditor_name,"city":city}
    meta={"mode":mode,"template":template,"entity_name":name,"short_name":entity["short_name"],"name_line2":entity["name_line2"],
          "activity_short":entity["activity_short"],"rc":rc,"auditor":auditor,"auditor_name":auditor_name,
          "fy":fy,"period_end":period_end,"sign_date":sign_date,"framework":fw,"framework_short":fws,
          "first_year":bool(first_year),"signatories":directors[:n_sig],"sig_words":SIG_WORDS.get(n_sig,str(n_sig)),
          "results_para":results_para,
          "ppe_para":"Movements in property, plant and equipment during the year are set out in the notes.",
          "frc_no":frc_no,"ican_stamp_no":ican_stamp_no,"stamp_image":stamp_image,"signature_image":signature_image,"total_pages":19}
    # tie-outs (reuse the 5-check logic)
    import afs_extract as ax
    checks=ax.tie_outs(soci,sofp_rows,scf,soce_rows)
    flags=[]
    if not rc: flags.append("RC number not provided (enter it on the engagement).")
    scf_end=ax._find(scf,"CASH","31") or ax._find(scf,"CASH","END")
    sofp_cash=ax._find(sofp_rows,"CASH","EQUIVALENTS")
    if abs((scf_end or 0)-(sofp_cash or 0))>1:
        flags.append("Cash flow closing cash (₦{:,.0f}) does not agree to balance-sheet cash (₦{:,.0f}); difference ₦{:,.0f} — review the source workbook.".format(scf_end or 0, sofp_cash or 0, (scf_end or 0)-(sofp_cash or 0)))
    flags.append("Entity details (name, RC, directors) were taken from the engagement form, not the workbook.")
    data={"meta":meta,"entity":entity,"soci":soci,"sofp":sofp_rows,"scf":scf,"soce":soce_rows,"notes":notes,
          "flags":flags,"tie_outs":[{"name":n,"pass":bool(ok)} for n,ok in checks]}
    return data
